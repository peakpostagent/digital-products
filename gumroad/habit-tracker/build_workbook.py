"""
Build script for "The 30-Day Habit Tracker & Goal Dashboard" Google Sheets workbook.
Generates a polished .xlsx file with 5 tabs using openpyxl.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from copy import copy

# ── Color Palette ──
NAVY       = "1E3A5F"
DARK_NAVY  = "152C4A"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F5F6FA"
MED_GRAY   = "E0E0E0"
DARK_GRAY  = "7F8C8D"
GREEN      = "27AE60"
LIGHT_GREEN= "E8F8F5"
RED        = "E74C3C"
LIGHT_RED  = "FDEDEC"
BLUE       = "3498DB"
ORANGE     = "F39C12"
TEAL       = "1ABC9C"
PURPLE     = "9B59B6"
PINK       = "E91E63"
AMBER      = "FF9800"
YELLOW_BG  = "FFF9C4"

# Category colors
CAT_COLORS = {
    "Health":       "27AE60",
    "Productivity": "3498DB",
    "Learning":     "9B59B6",
    "Mindfulness":  "1ABC9C",
    "Fitness":      "E67E22",
    "Creative":     "E91E63",
    "Social":       "F39C12",
    "Custom":       "7F8C8D",
}

# ── Style helpers ──
thin_border = Border(
    left=Side(style="thin", color=MED_GRAY),
    right=Side(style="thin", color=MED_GRAY),
    top=Side(style="thin", color=MED_GRAY),
    bottom=Side(style="thin", color=MED_GRAY),
)

def header_font(size=12, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=11, bold=False, color="333333", italic=False):
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

def navy_fill():
    return PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

def white_fill():
    return PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

def light_fill():
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

def green_fill():
    return PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")

def red_fill():
    return PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")

def color_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border


# ── Sample data for first 5 days ──
SAMPLE_HABITS = [
    ("Drink 8 glasses of water", "Health", "Drink 8 full glasses throughout the day", "Stay hydrated for energy & focus"),
    ("Read for 20 minutes", "Learning", "Read a book or article for 20 min", "Expand knowledge & reduce stress"),
    ("Meditate for 5 minutes", "Mindfulness", "Sit quietly and focus on breathing", "Reduce anxiety & improve clarity"),
    ("Exercise 30 minutes", "Fitness", "Any physical activity for 30 min", "Build strength & boost mood"),
    ("Write in journal", "Creative", "Write at least 200 words", "Process thoughts & track growth"),
    ("No phone first hour", "Productivity", "Avoid phone for 1st hour after waking", "Start the day with intention"),
    ("Practice a skill", "Learning", "Spend 15 min on a new skill", "Continuous improvement compounds"),
    ("Connect with someone", "Social", "Call, text, or meet a friend/family", "Nurture meaningful relationships"),
]

# Sample tracking data (rows=habits, cols=days 1-5)
# 1=done, 0=missed
SAMPLE_DATA = [
    [1, 1, 1, 0, 1],
    [1, 1, 1, 1, 1],
    [1, 0, 1, 1, 1],
    [0, 1, 1, 1, 0],
    [1, 1, 0, 1, 1],
    [1, 1, 1, 1, 0],
    [0, 1, 1, 0, 1],
    [1, 1, 1, 1, 1],
]


def build_setup_tab(wb):
    ws = wb.active
    ws.title = "Setup"
    ws.sheet_properties.tabColor = NAVY

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 5

    # Freeze top rows
    ws.freeze_panes = "A5"

    # ── Title banner ──
    ws.merge_cells("B1:E1")
    title_cell = ws.cell(row=1, column=2, value="THE 30-DAY HABIT TRACKER & GOAL DASHBOARD")
    title_cell.font = header_font(18)
    title_cell.fill = navy_fill()
    title_cell.alignment = center_align()
    style_range(ws, 1, 2, 5, font=header_font(18), fill=navy_fill(), alignment=center_align())

    ws.merge_cells("B2:E2")
    sub = ws.cell(row=2, column=2, value="Track your habits, crush your goals, and see your progress.")
    sub.font = Font(name="Calibri", size=11, italic=True, color=WHITE)
    sub.fill = color_fill(DARK_NAVY)
    sub.alignment = center_align()
    style_range(ws, 2, 2, 5, font=Font(name="Calibri", size=11, italic=True, color=WHITE), fill=color_fill(DARK_NAVY), alignment=center_align())

    # Row 3 spacer
    ws.row_dimensions[3].height = 8

    # ── Your Info ──
    row = 4
    ws.merge_cells(f"B{row}:E{row}")
    ws.cell(row=row, column=2, value="YOUR INFO").font = header_font(13, color=NAVY)
    ws.cell(row=row, column=2).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 5, fill=color_fill(LIGHT_GRAY))

    row = 5
    ws.cell(row=row, column=2, value="Your Name:").font = body_font(bold=True)
    ws.cell(row=row, column=3, value="").font = body_font()
    ws.cell(row=row, column=3).border = Border(bottom=Side(style="thin", color=NAVY))

    row = 6
    ws.cell(row=row, column=2, value="Start Date:").font = body_font(bold=True)
    ws.cell(row=row, column=3, value="2026-03-18").font = body_font()
    ws.cell(row=row, column=3).number_format = "YYYY-MM-DD"
    ws.cell(row=row, column=3).border = Border(bottom=Side(style="thin", color=NAVY))

    row = 7
    ws.cell(row=row, column=2, value="End Date (auto):").font = body_font(bold=True)
    ws.cell(row=row, column=3).font = body_font()
    ws.cell(row=row, column=3, value="=C6+29")
    ws.cell(row=row, column=3).number_format = "YYYY-MM-DD"
    ws.cell(row=row, column=3).fill = color_fill(YELLOW_BG)

    # Spacer
    ws.row_dimensions[8].height = 8

    # ── Define Your Habits ──
    row = 9
    ws.merge_cells(f"B{row}:E{row}")
    ws.cell(row=row, column=2, value="DEFINE YOUR HABITS (UP TO 8)").font = header_font(13, color=NAVY)
    ws.cell(row=row, column=2).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 5, fill=color_fill(LIGHT_GRAY))

    row = 10
    headers = ["Habit Name", "Category", "Daily Goal Description", "Why It Matters"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=2+i, value=h)
        cell.font = header_font(11, color=WHITE)
        cell.fill = navy_fill()
        cell.alignment = center_align()
        cell.border = thin_border

    # Habit rows with sample data
    from openpyxl.worksheet.datavalidation import DataValidation
    cat_dv = DataValidation(
        type="list",
        formula1='"Health,Productivity,Learning,Mindfulness,Fitness,Creative,Social,Custom"',
        allow_blank=True
    )
    cat_dv.error = "Please select a category from the dropdown."
    cat_dv.errorTitle = "Invalid Category"
    ws.add_data_validation(cat_dv)

    for i in range(8):
        row = 11 + i
        ws.row_dimensions[row].height = 30
        bg = white_fill() if i % 2 == 0 else light_fill()

        # Number label
        ws.cell(row=row, column=1, value=i+1).font = body_font(bold=True, color=DARK_GRAY)
        ws.cell(row=row, column=1).alignment = center_align()

        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.fill = bg
            cell.border = thin_border
            cell.alignment = left_align()
            cell.font = body_font()

        # Fill sample data
        if i < len(SAMPLE_HABITS):
            name, cat, goal, why = SAMPLE_HABITS[i]
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=cat)
            ws.cell(row=row, column=4, value=goal)
            ws.cell(row=row, column=5, value=why)

        # Category dropdown
        cat_dv.add(ws.cell(row=row, column=3))

    # ── Category Color Guide ──
    row = 21
    ws.row_dimensions[row].height = 8
    row = 22
    ws.merge_cells(f"B{row}:E{row}")
    ws.cell(row=row, column=2, value="CATEGORY COLOR GUIDE").font = header_font(13, color=NAVY)
    ws.cell(row=row, column=2).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 5, fill=color_fill(LIGHT_GRAY))

    row = 23
    col = 2
    for cat_name, cat_color in CAT_COLORS.items():
        r = row + (list(CAT_COLORS.keys()).index(cat_name) // 4)
        c = col + (list(CAT_COLORS.keys()).index(cat_name) % 4)
        cell = ws.cell(row=r, column=c, value=cat_name)
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.fill = color_fill(cat_color)
        cell.alignment = center_align()
        cell.border = thin_border

    # ── Instructions ──
    row = 26
    ws.row_dimensions[row].height = 8
    row = 27
    ws.merge_cells(f"B{row}:E{row}")
    ws.cell(row=row, column=2, value="HOW TO USE THIS WORKBOOK").font = header_font(13, color=NAVY)
    ws.cell(row=row, column=2).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 5, fill=color_fill(LIGHT_GRAY))

    instructions = [
        "1. SETUP: Fill in your name, start date, and define up to 8 habits above.",
        "2. TRACK: Go to the '30-Day Tracker' tab. Enter 1 for completed, 0 for missed. Leave future days blank.",
        "3. DASHBOARD: Check the 'Dashboard' tab anytime to see your progress, streaks, and completion rates.",
        "4. REVIEW: At the end of each week, use the 'Weekly Review' tab to reflect on wins and struggles.",
        "5. IDEAS: Browse the 'Habit Ideas Bank' tab for 50+ habit suggestions if you need inspiration.",
        "",
        "TIP: Start with just 3-4 habits. You can always add more once you build consistency!",
        "TIP: Open this in the Google Sheets mobile app to track habits on the go.",
        "TIP: Share the link with a friend for accountability.",
    ]
    for i, inst in enumerate(instructions):
        r = 28 + i
        ws.merge_cells(f"B{r}:E{r}")
        ws.cell(row=r, column=2, value=inst).font = body_font(color="333333" if inst.startswith(("1","2","3","4","5")) else DARK_GRAY)
        ws.cell(row=r, column=2).alignment = left_align()

    return ws


def build_tracker_tab(wb):
    ws = wb.create_sheet("30-Day Tracker")
    ws.sheet_properties.tabColor = GREEN

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    for i in range(3, 33):  # Days 1-30
        ws.column_dimensions[get_column_letter(i)].width = 5
    ws.column_dimensions[get_column_letter(33)].width = 12  # Streak
    ws.column_dimensions[get_column_letter(34)].width = 12  # Best Streak
    ws.column_dimensions[get_column_letter(35)].width = 14  # Completion %

    # Freeze panes
    ws.freeze_panes = "C3"

    # ── Title ──
    ws.merge_cells("B1:AI1")
    ws.cell(row=1, column=2, value="30-DAY HABIT TRACKER").font = header_font(16)
    ws.cell(row=1, column=2).fill = navy_fill()
    ws.cell(row=1, column=2).alignment = center_align()
    style_range(ws, 1, 2, 35, fill=navy_fill(), font=header_font(16), alignment=center_align())

    # ── Day headers ──
    ws.cell(row=2, column=2, value="HABIT").font = header_font(11, color=WHITE)
    ws.cell(row=2, column=2).fill = color_fill(DARK_NAVY)
    ws.cell(row=2, column=2).alignment = center_align()
    ws.cell(row=2, column=2).border = thin_border

    for day in range(1, 31):
        col = day + 2
        cell = ws.cell(row=2, column=col, value=day)
        cell.font = header_font(10, color=WHITE)
        cell.fill = color_fill(DARK_NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    # Stats headers
    stats_headers = [("Current\nStreak", 33), ("Best\nStreak", 34), ("Completion\n%", 35)]
    for label, col in stats_headers:
        cell = ws.cell(row=2, column=col, value=label)
        cell.font = header_font(9, color=WHITE)
        cell.fill = color_fill(NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    # ── Habit rows (rows 3-10) ──
    for i in range(8):
        row = 3 + i
        ws.row_dimensions[row].height = 28
        bg = white_fill() if i % 2 == 0 else light_fill()

        # Row number
        ws.cell(row=row, column=1, value=i+1).font = body_font(bold=True, color=DARK_GRAY)
        ws.cell(row=row, column=1).alignment = center_align()

        # Habit name (linked to Setup)
        habit_cell = ws.cell(row=row, column=2)
        habit_cell.value = f"=Setup!B{11+i}"
        habit_cell.font = body_font(bold=True)
        habit_cell.fill = bg
        habit_cell.alignment = left_align()
        habit_cell.border = thin_border

        # Day cells
        for day in range(1, 31):
            col = day + 2
            cell = ws.cell(row=row, column=col)
            cell.alignment = center_align()
            cell.border = thin_border
            cell.font = body_font()

            # Fill sample data for first 5 days
            if day <= 5 and i < len(SAMPLE_DATA):
                cell.value = SAMPLE_DATA[i][day - 1]

        # Current Streak formula
        # Count consecutive 1s from the rightmost filled cell going left
        streak_cell = ws.cell(row=row, column=33)
        # Build a formula that counts the current streak from the end
        # We'll use a simpler approach: SUMPRODUCT to find consecutive 1s from the last entry
        streak_formula = (
            f'=IF(COUNTA(C{row}:AF{row})=0,0,'
            f'LET(n,COUNTA(C{row}:AF{row}),'
            f'SUMPRODUCT(--('
            f'MMULT(--({row}={row}),SEQUENCE(n,1,1,0)^0)='
            f'MMULT(--({row}={row}),SEQUENCE(n,1,1,0)^0))))'
        )
        # Simpler streak: just count from the right
        # Actually let's use a straightforward formula
        streak_cell.value = (
            f'=IF(COUNTA(C{row}:AF{row})=0,0,'
            f'IF(INDIRECT("R{row}C"&(2+COUNTA(C{row}:AF{row})),FALSE)=1,'
            f'COUNTA(C{row}:AF{row})-IFERROR(MATCH(0,INDEX(C{row}:AF{row},SEQUENCE(COUNTA(C{row}:AF{row}),1,COUNTA(C{row}:AF{row}),-1)),0),COUNTA(C{row}:AF{row}))+1,'
            f'0))'
        )
        streak_cell.font = body_font(bold=True, color=GREEN)
        streak_cell.alignment = center_align()
        streak_cell.border = thin_border

        # Best Streak formula (simplified - max consecutive 1s)
        best_cell = ws.cell(row=row, column=34)
        best_cell.value = f'=IF(COUNTA(C{row}:AF{row})=0,0,MAX(FREQUENCY(IF(C{row}:AF{row}=1,COLUMN(C{row}:AF{row})),IF(C{row}:AF{row}<>1,COLUMN(C{row}:AF{row})))))'
        best_cell.font = body_font(bold=True, color=BLUE)
        best_cell.alignment = center_align()
        best_cell.border = thin_border

        # Completion % formula
        comp_cell = ws.cell(row=row, column=35)
        comp_cell.value = f'=IF(COUNTA(C{row}:AF{row})=0,"",COUNTIF(C{row}:AF{row},1)/COUNTA(C{row}:AF{row}))'
        comp_cell.font = body_font(bold=True, color=NAVY)
        comp_cell.number_format = "0%"
        comp_cell.alignment = center_align()
        comp_cell.border = thin_border

    # ── Daily Completion % row ──
    row = 12
    ws.row_dimensions[row].height = 8
    row = 13
    ws.cell(row=row, column=2, value="DAILY COMPLETION %").font = header_font(10, color=WHITE)
    ws.cell(row=row, column=2).fill = color_fill(NAVY)
    ws.cell(row=row, column=2).alignment = center_align()
    ws.cell(row=row, column=2).border = thin_border

    for day in range(1, 31):
        col = day + 2
        cell = ws.cell(row=row, column=col)
        cell.value = f'=IF(COUNTA(C3:C10)=0,"",COUNTIF({get_column_letter(col)}3:{get_column_letter(col)}10,1)/COUNTA({get_column_letter(col)}3:{get_column_letter(col)}10))'
        cell.number_format = "0%"
        cell.font = body_font(bold=True, color=NAVY)
        cell.fill = color_fill(YELLOW_BG)
        cell.alignment = center_align()
        cell.border = thin_border

    # ── Daily Notes row ──
    row = 15
    ws.cell(row=row, column=2, value="DAILY NOTES / REFLECTIONS").font = header_font(10, color=WHITE)
    ws.cell(row=row, column=2).fill = color_fill(NAVY)
    ws.cell(row=row, column=2).alignment = center_align()
    ws.cell(row=row, column=2).border = thin_border

    row = 16
    ws.row_dimensions[row].height = 50
    for day in range(1, 31):
        col = day + 2
        cell = ws.cell(row=row, column=col)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border = thin_border
        cell.font = body_font(size=9)

    # Sample notes
    ws.cell(row=16, column=3, value="Great start! Missed the gym but stayed on track with everything else.")
    ws.cell(row=16, column=4, value="Much better today. Got all my reading done.")
    ws.cell(row=16, column=5, value="Tired but pushed through meditation.")

    # ── Conditional formatting ──
    # Green for 1 (done)
    green_rule = CellIsRule(
        operator="equal", formula=["1"],
        fill=PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
        font=Font(color="1B5E20", bold=True)
    )
    # Red for 0 (missed)
    red_rule = CellIsRule(
        operator="equal", formula=["0"],
        fill=PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
        font=Font(color="B71C1C", bold=True)
    )

    data_range = "C3:AF10"
    ws.conditional_formatting.add(data_range, green_rule)
    ws.conditional_formatting.add(data_range, red_rule)

    return ws


def build_dashboard_tab(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = BLUE

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 25

    # ── Title ──
    ws.merge_cells("B1:H1")
    ws.cell(row=1, column=2, value="PROGRESS DASHBOARD").font = header_font(18)
    ws.cell(row=1, column=2).fill = navy_fill()
    ws.cell(row=1, column=2).alignment = center_align()
    style_range(ws, 1, 2, 8, fill=navy_fill(), font=header_font(18), alignment=center_align())

    ws.merge_cells("B2:H2")
    ws.cell(row=2, column=2, value="Your habit-building journey at a glance").font = Font(name="Calibri", size=11, italic=True, color=WHITE)
    ws.cell(row=2, column=2).fill = color_fill(DARK_NAVY)
    ws.cell(row=2, column=2).alignment = center_align()
    style_range(ws, 2, 2, 8, fill=color_fill(DARK_NAVY), alignment=center_align())

    # ── Big Stats Row ──
    row = 4
    ws.merge_cells(f"B{row}:C{row}")
    ws.cell(row=row, column=2, value="OVERALL COMPLETION RATE").font = header_font(11, color=NAVY)
    ws.cell(row=row, column=2).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=2).alignment = center_align()
    style_range(ws, row, 2, 3, fill=color_fill(LIGHT_GRAY))

    ws.merge_cells(f"D{row}:E{row}")
    ws.cell(row=row, column=4, value="DAYS COMPLETED").font = header_font(11, color=NAVY)
    ws.cell(row=row, column=4).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=4).alignment = center_align()
    style_range(ws, row, 4, 5, fill=color_fill(LIGHT_GRAY))

    ws.merge_cells(f"F{row}:G{row}")
    ws.cell(row=row, column=6, value="ACTIVE HABITS").font = header_font(11, color=NAVY)
    ws.cell(row=row, column=6).fill = color_fill(LIGHT_GRAY)
    ws.cell(row=row, column=6).alignment = center_align()
    style_range(ws, row, 6, 7, fill=color_fill(LIGHT_GRAY))

    row = 5
    ws.row_dimensions[row].height = 55

    # Overall completion rate
    ws.merge_cells(f"B{row}:C{row}")
    cell = ws.cell(row=row, column=2)
    cell.value = '=IFERROR(COUNTIF(\'30-Day Tracker\'!C3:AF10,1)/COUNTA(\'30-Day Tracker\'!C3:AF10),0)'
    cell.font = Font(name="Calibri", size=36, bold=True, color=GREEN)
    cell.number_format = "0%"
    cell.alignment = center_align()
    cell.fill = white_fill()
    cell.border = thin_border
    ws.cell(row=row, column=3).border = thin_border

    # Days completed
    ws.merge_cells(f"D{row}:E{row}")
    cell = ws.cell(row=row, column=4)
    cell.value = '=IFERROR(MAX(COUNTA(\'30-Day Tracker\'!C3:AF3),COUNTA(\'30-Day Tracker\'!C4:AF4),COUNTA(\'30-Day Tracker\'!C5:AF5),COUNTA(\'30-Day Tracker\'!C6:AF6),COUNTA(\'30-Day Tracker\'!C7:AF7),COUNTA(\'30-Day Tracker\'!C8:AF8),COUNTA(\'30-Day Tracker\'!C9:AF9),COUNTA(\'30-Day Tracker\'!C10:AF10)),0)&" of 30"'
    cell.font = Font(name="Calibri", size=28, bold=True, color=NAVY)
    cell.alignment = center_align()
    cell.fill = white_fill()
    cell.border = thin_border

    # Active habits
    ws.merge_cells(f"F{row}:G{row}")
    cell = ws.cell(row=row, column=6)
    cell.value = '=COUNTA(Setup!B11:B18)'
    cell.font = Font(name="Calibri", size=36, bold=True, color=BLUE)
    cell.alignment = center_align()
    cell.fill = white_fill()
    cell.border = thin_border

    # ── Per-Habit Completion Rates ──
    row = 7
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(row=row, column=2, value="PER-HABIT COMPLETION RATES").font = header_font(12, color=WHITE)
    ws.cell(row=row, column=2).fill = navy_fill()
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 8, fill=navy_fill())

    row = 8
    headers = ["Habit", "Completion %", "Current Streak", "Best Streak", "Status", "Visual Progress"]
    cols = [2, 3, 4, 5, 6, 7]
    for h, c in zip(headers, cols):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font(10, color=WHITE)
        cell.fill = color_fill(DARK_NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    for i in range(8):
        row = 9 + i
        ws.row_dimensions[row].height = 25
        bg = white_fill() if i % 2 == 0 else light_fill()
        tracker_row = 3 + i

        # Habit name
        cell = ws.cell(row=row, column=2, value=f"=Setup!B{11+i}")
        cell.font = body_font(bold=True)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = left_align()

        # Completion %
        cell = ws.cell(row=row, column=3)
        cell.value = f"='30-Day Tracker'!AI{tracker_row}"
        cell.number_format = "0%"
        cell.font = body_font(bold=True, color=NAVY)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Current streak
        cell = ws.cell(row=row, column=4)
        cell.value = f"='30-Day Tracker'!AG{tracker_row}"
        cell.font = body_font(bold=True, color=GREEN)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Best streak
        cell = ws.cell(row=row, column=5)
        cell.value = f"='30-Day Tracker'!AH{tracker_row}"
        cell.font = body_font(bold=True, color=BLUE)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Status (emoji-like text)
        cell = ws.cell(row=row, column=6)
        cell.value = f'=IF(Setup!B{11+i}="","",IF(\'30-Day Tracker\'!AI{tracker_row}>=0.8,"Excellent",IF(\'30-Day Tracker\'!AI{tracker_row}>=0.6,"Good",IF(\'30-Day Tracker\'!AI{tracker_row}>=0.4,"Fair","Needs Work"))))'
        cell.font = body_font(bold=True)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Visual progress bar using REPT
        cell = ws.cell(row=row, column=7)
        cell.value = f'=IF(Setup!B{11+i}="","",REPT("|",ROUND(IFERROR(\'30-Day Tracker\'!AI{tracker_row},0)*20,0))&REPT(".",20-ROUND(IFERROR(\'30-Day Tracker\'!AI{tracker_row},0)*20,0)))'
        cell.font = Font(name="Consolas", size=10, bold=True, color=GREEN)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = left_align()

    # ── Weekly Breakdown ──
    row = 19
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(row=row, column=2, value="WEEKLY BREAKDOWN").font = header_font(12, color=WHITE)
    ws.cell(row=row, column=2).fill = navy_fill()
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 8, fill=navy_fill())

    row = 20
    week_headers = ["Week", "Days", "Completion Rate", "Visual"]
    for i, h in enumerate(week_headers):
        cell = ws.cell(row=row, column=2+i, value=h)
        cell.font = header_font(10, color=WHITE)
        cell.fill = color_fill(DARK_NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    week_ranges = [
        ("Week 1", "C", "I"),   # Days 1-7
        ("Week 2", "J", "P"),   # Days 8-14
        ("Week 3", "Q", "W"),   # Days 15-21
        ("Week 4", "X", "AF"),  # Days 22-30
    ]

    for i, (label, start_col, end_col) in enumerate(week_ranges):
        row = 21 + i
        bg = white_fill() if i % 2 == 0 else light_fill()

        ws.cell(row=row, column=2, value=label).font = body_font(bold=True)
        ws.cell(row=row, column=2).fill = bg
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=2).alignment = center_align()

        days = f"Days {i*7+1}-{min((i+1)*7, 30)}" if i < 3 else "Days 22-30"
        ws.cell(row=row, column=3, value=days).font = body_font()
        ws.cell(row=row, column=3).fill = bg
        ws.cell(row=row, column=3).border = thin_border
        ws.cell(row=row, column=3).alignment = center_align()

        # Weekly completion rate
        cell = ws.cell(row=row, column=4)
        cell.value = f"=IFERROR(COUNTIF('30-Day Tracker'!{start_col}3:{end_col}10,1)/COUNTA('30-Day Tracker'!{start_col}3:{end_col}10),0)"
        cell.number_format = "0%"
        cell.font = body_font(bold=True, color=NAVY)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Visual bar
        cell = ws.cell(row=row, column=5)
        cell.value = f'=REPT("|",ROUND(D{row}*20,0))&REPT(".",20-ROUND(D{row}*20,0))'
        cell.font = Font(name="Consolas", size=10, bold=True, color=TEAL)
        cell.fill = bg
        cell.border = thin_border

    # ── Category Performance ──
    row = 27
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(row=row, column=2, value="CATEGORY PERFORMANCE").font = header_font(12, color=WHITE)
    ws.cell(row=row, column=2).fill = navy_fill()
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 8, fill=navy_fill())

    row = 28
    for i, h in enumerate(["Category", "Avg Completion", "Visual"]):
        cell = ws.cell(row=row, column=2+i, value=h)
        cell.font = header_font(10, color=WHITE)
        cell.fill = color_fill(DARK_NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    for i, cat in enumerate(CAT_COLORS.keys()):
        row = 29 + i
        bg = white_fill() if i % 2 == 0 else light_fill()

        cell = ws.cell(row=row, column=2, value=cat)
        cell.font = body_font(bold=True)
        cell.fill = color_fill(CAT_COLORS[cat])
        cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
        cell.alignment = center_align()
        cell.border = thin_border

        # Average completion for this category
        cell = ws.cell(row=row, column=3)
        cell.value = f'=IFERROR(AVERAGEIF(Setup!C11:C18,"{cat}",\'30-Day Tracker\'!AI3:AI10),"")'
        cell.number_format = "0%"
        cell.font = body_font(bold=True)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Visual bar
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(C{row}="","",REPT("|",ROUND(C{row}*20,0))&REPT(".",20-ROUND(C{row}*20,0)))'
        cell.font = Font(name="Consolas", size=10, bold=True, color=CAT_COLORS[cat])
        cell.fill = bg
        cell.border = thin_border

    # ── Milestone Tracker ──
    row = 39
    ws.merge_cells(f"B{row}:H{row}")
    ws.cell(row=row, column=2, value="MILESTONE TRACKER").font = header_font(12, color=WHITE)
    ws.cell(row=row, column=2).fill = navy_fill()
    ws.cell(row=row, column=2).alignment = left_align()
    style_range(ws, row, 2, 8, fill=navy_fill())

    milestones = [
        (7, "1 WEEK", "You've built the foundation!"),
        (14, "2 WEEKS", "You're forming real habits!"),
        (21, "3 WEEKS", "Science says this is when habits stick!"),
        (30, "30 DAYS", "CONGRATULATIONS! You did it!"),
    ]

    row = 40
    for i, (days, label, msg) in enumerate(milestones):
        r = row + i
        ws.row_dimensions[r].height = 30
        bg = white_fill() if i % 2 == 0 else light_fill()

        cell = ws.cell(row=r, column=2, value=label)
        cell.font = Font(name="Calibri", size=13, bold=True, color=NAVY)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        # Status
        cell = ws.cell(row=r, column=3)
        cell.value = f'=IF(IFERROR(MAX(COUNTA(\'30-Day Tracker\'!C3:AF3),COUNTA(\'30-Day Tracker\'!C4:AF4),COUNTA(\'30-Day Tracker\'!C5:AF5),COUNTA(\'30-Day Tracker\'!C6:AF6),COUNTA(\'30-Day Tracker\'!C7:AF7),COUNTA(\'30-Day Tracker\'!C8:AF8),COUNTA(\'30-Day Tracker\'!C9:AF9),COUNTA(\'30-Day Tracker\'!C10:AF10)),0)>={days},"ACHIEVED","In Progress")'
        cell.font = body_font(bold=True, color=GREEN)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        ws.merge_cells(f"D{r}:H{r}")
        cell = ws.cell(row=r, column=4, value=msg)
        cell.font = body_font(italic=True, color=DARK_GRAY)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = left_align()

    return ws


def build_weekly_review_tab(wb):
    ws = wb.create_sheet("Weekly Review")
    ws.sheet_properties.tabColor = PURPLE

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 5

    # Title
    ws.merge_cells("B1:C1")
    ws.cell(row=1, column=2, value="WEEKLY REVIEW & REFLECTION").font = header_font(16)
    ws.cell(row=1, column=2).fill = navy_fill()
    ws.cell(row=1, column=2).alignment = center_align()
    style_range(ws, 1, 2, 3, fill=navy_fill(), font=header_font(16), alignment=center_align())

    ws.merge_cells("B2:C2")
    ws.cell(row=2, column=2, value="Take 10 minutes each week to reflect on your progress").font = Font(name="Calibri", size=11, italic=True, color=WHITE)
    ws.cell(row=2, column=2).fill = color_fill(DARK_NAVY)
    ws.cell(row=2, column=2).alignment = center_align()
    style_range(ws, 2, 2, 3, fill=color_fill(DARK_NAVY), alignment=center_align())

    week_col_ranges = [
        ("C", "I"),   # Week 1: Days 1-7
        ("J", "P"),   # Week 2: Days 8-14
        ("Q", "W"),   # Week 3: Days 15-21
        ("X", "AF"),  # Week 4: Days 22-30
    ]

    current_row = 4
    for w in range(4):
        week_num = w + 1
        start_col, end_col = week_col_ranges[w]
        day_start = w * 7 + 1
        day_end = min((w + 1) * 7, 30)

        # Week header
        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws.cell(row=current_row, column=2, value=f"WEEK {week_num} REVIEW (Days {day_start}-{day_end})").font = header_font(13, color=WHITE)
        ws.cell(row=current_row, column=2).fill = navy_fill()
        ws.cell(row=current_row, column=2).alignment = left_align()
        style_range(ws, current_row, 2, 3, fill=navy_fill())
        current_row += 1

        # Auto stats
        ws.cell(row=current_row, column=2, value="Week Completion Rate:").font = body_font(bold=True, color=NAVY)
        ws.cell(row=current_row, column=2).fill = color_fill(LIGHT_GRAY)
        ws.cell(row=current_row, column=2).border = thin_border
        cell = ws.cell(row=current_row, column=3)
        cell.value = f"=IFERROR(COUNTIF('30-Day Tracker'!{start_col}3:{end_col}10,1)/COUNTA('30-Day Tracker'!{start_col}3:{end_col}10),0)"
        cell.number_format = "0%"
        cell.font = Font(name="Calibri", size=16, bold=True, color=GREEN)
        cell.fill = color_fill(LIGHT_GRAY)
        cell.border = thin_border
        cell.alignment = center_align()
        current_row += 1

        ws.cell(row=current_row, column=2, value="Habits Completed:").font = body_font(bold=True, color=NAVY)
        ws.cell(row=current_row, column=2).fill = color_fill(LIGHT_GRAY)
        ws.cell(row=current_row, column=2).border = thin_border
        cell = ws.cell(row=current_row, column=3)
        cell.value = f"=COUNTIF('30-Day Tracker'!{start_col}3:{end_col}10,1)&\" / \"&COUNTA('30-Day Tracker'!{start_col}3:{end_col}10)"
        cell.font = body_font(bold=True)
        cell.fill = color_fill(LIGHT_GRAY)
        cell.border = thin_border
        cell.alignment = center_align()
        current_row += 1

        # Reflection prompts
        prompts = [
            ("What went well this week?", "What habits did you nail? What felt easy or natural?"),
            ("What was challenging?", "Which habits did you miss? What got in the way?"),
            ("What will you adjust?", "Any habits to modify, swap, or approach differently?"),
            (f"Focus for {'Next Week' if week_num < 4 else 'Going Forward'}:", "What one thing will make the biggest difference?"),
        ]

        for prompt_title, hint in prompts:
            ws.cell(row=current_row, column=2, value=prompt_title).font = body_font(bold=True, color=NAVY)
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).fill = white_fill()
            current_row += 1

            ws.row_dimensions[current_row].height = 60
            cell = ws.cell(row=current_row, column=2, value=hint)
            cell.font = body_font(size=10, color=DARK_GRAY)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border

            ws.merge_cells(f"B{current_row}:C{current_row}")
            current_row += 1

        # Spacer
        current_row += 1

    return ws


def build_habit_ideas_tab(wb):
    ws = wb.create_sheet("Habit Ideas Bank")
    ws.sheet_properties.tabColor = TEAL

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Title
    ws.merge_cells("B1:E1")
    ws.cell(row=1, column=2, value="HABIT IDEAS BANK").font = header_font(16)
    ws.cell(row=1, column=2).fill = navy_fill()
    ws.cell(row=1, column=2).alignment = center_align()
    style_range(ws, 1, 2, 5, fill=navy_fill(), font=header_font(16), alignment=center_align())

    ws.merge_cells("B2:E2")
    ws.cell(row=2, column=2, value="Browse 50+ ideas to find habits that work for you. Start with 3-4 Easy ones!").font = Font(name="Calibri", size=11, italic=True, color=WHITE)
    ws.cell(row=2, column=2).fill = color_fill(DARK_NAVY)
    ws.cell(row=2, column=2).alignment = center_align()
    style_range(ws, 2, 2, 5, fill=color_fill(DARK_NAVY), alignment=center_align())

    # Column headers
    row = 3
    for i, h in enumerate(["Habit Idea", "Category", "Difficulty", "Time Needed"]):
        cell = ws.cell(row=row, column=2+i, value=h)
        cell.font = header_font(11, color=WHITE)
        cell.fill = color_fill(DARK_NAVY)
        cell.alignment = center_align()
        cell.border = thin_border

    # Habit ideas data
    habits = [
        # Health
        ("Drink 8 glasses of water", "Health", "Easy", "Throughout day"),
        ("Sleep by 10:00 PM", "Health", "Medium", "N/A"),
        ("Eat a serving of vegetables", "Health", "Easy", "At meals"),
        ("No sugar or processed snacks", "Health", "Hard", "Throughout day"),
        ("Walk for 10 minutes", "Health", "Easy", "10 min"),
        ("Stretch for 5 minutes", "Health", "Easy", "5 min"),
        ("Take vitamins/supplements", "Health", "Easy", "1 min"),
        ("No eating after 8 PM", "Health", "Medium", "N/A"),
        ("Cook a healthy meal", "Health", "Medium", "30 min"),
        # Productivity
        ("Plan tomorrow tonight", "Productivity", "Easy", "5 min"),
        ("No phone for first hour", "Productivity", "Medium", "60 min"),
        ("Time block your day", "Productivity", "Medium", "10 min"),
        ("Inbox zero", "Productivity", "Medium", "15 min"),
        ("Deep work for 1 hour", "Productivity", "Hard", "60 min"),
        ("Review daily goals", "Productivity", "Easy", "5 min"),
        ("Single-task (no multitasking)", "Productivity", "Hard", "Throughout day"),
        ("Complete 1 important task first", "Productivity", "Medium", "30-60 min"),
        ("Batch similar tasks together", "Productivity", "Medium", "Throughout day"),
        # Learning
        ("Read 10 pages of a book", "Learning", "Easy", "15 min"),
        ("Learn 5 vocabulary words", "Learning", "Easy", "10 min"),
        ("Watch an educational video", "Learning", "Easy", "15 min"),
        ("Practice a new skill for 15 min", "Learning", "Medium", "15 min"),
        ("Listen to a podcast episode", "Learning", "Easy", "20-30 min"),
        ("Take notes on what you learned", "Learning", "Easy", "5 min"),
        ("Teach someone something", "Learning", "Medium", "10 min"),
        ("Complete 1 online course lesson", "Learning", "Medium", "20 min"),
        # Mindfulness
        ("Meditate for 5 minutes", "Mindfulness", "Easy", "5 min"),
        ("Write 3 things you're grateful for", "Mindfulness", "Easy", "3 min"),
        ("Digital detox for 1 hour", "Mindfulness", "Medium", "60 min"),
        ("Deep breathing exercise", "Mindfulness", "Easy", "3 min"),
        ("Body scan relaxation", "Mindfulness", "Easy", "10 min"),
        ("Mindful eating (no screens)", "Mindfulness", "Medium", "15 min"),
        ("Spend 10 min in nature", "Mindfulness", "Easy", "10 min"),
        ("Evening wind-down routine", "Mindfulness", "Medium", "15 min"),
        # Fitness
        ("Do 20 pushups", "Fitness", "Medium", "3 min"),
        ("Walk 10,000 steps", "Fitness", "Medium", "60-90 min"),
        ("10 minutes of yoga", "Fitness", "Easy", "10 min"),
        ("Run or jog 1 mile", "Fitness", "Medium", "10-15 min"),
        ("Hold plank for 1 minute", "Fitness", "Medium", "2 min"),
        ("Do 50 squats", "Fitness", "Medium", "5 min"),
        ("Take the stairs (avoid elevators)", "Fitness", "Easy", "Throughout day"),
        ("Full body workout", "Fitness", "Hard", "30-45 min"),
        # Creative
        ("Write 200 words", "Creative", "Easy", "10 min"),
        ("Sketch or draw something", "Creative", "Easy", "10 min"),
        ("Play an instrument for 10 min", "Creative", "Medium", "10 min"),
        ("Photograph something interesting", "Creative", "Easy", "5 min"),
        ("Brainstorm 5 ideas", "Creative", "Easy", "5 min"),
        ("Work on a creative project", "Creative", "Medium", "20 min"),
        ("Color or paint for relaxation", "Creative", "Easy", "15 min"),
        # Social
        ("Call or text a friend", "Social", "Easy", "5-10 min"),
        ("Give someone a genuine compliment", "Social", "Easy", "1 min"),
        ("Write a thank you note", "Social", "Easy", "5 min"),
        ("Help someone with something", "Social", "Easy", "10 min"),
        ("Have a tech-free conversation", "Social", "Medium", "15 min"),
        ("Send an encouraging message", "Social", "Easy", "3 min"),
        ("Eat a meal with someone", "Social", "Easy", "30 min"),
        ("Practice active listening", "Social", "Medium", "Throughout day"),
    ]

    current_cat = ""
    row = 4
    for habit, cat, diff, time_needed in habits:
        # Category separator
        if cat != current_cat:
            if current_cat != "":
                row += 1  # spacer
            current_cat = cat

        bg = white_fill() if (row - 4) % 2 == 0 else light_fill()

        cell = ws.cell(row=row, column=2, value=habit)
        cell.font = body_font()
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = left_align()

        cell = ws.cell(row=row, column=3, value=cat)
        cat_color = CAT_COLORS.get(cat, DARK_GRAY)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = color_fill(cat_color)
        cell.border = thin_border
        cell.alignment = center_align()

        # Difficulty with color coding
        cell = ws.cell(row=row, column=4, value=diff)
        diff_colors = {"Easy": GREEN, "Medium": ORANGE, "Hard": RED}
        cell.font = Font(name="Calibri", size=10, bold=True, color=diff_colors.get(diff, DARK_GRAY))
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        cell = ws.cell(row=row, column=5, value=time_needed)
        cell.font = body_font(size=10)
        cell.fill = bg
        cell.border = thin_border
        cell.alignment = center_align()

        row += 1

    return ws


def main():
    wb = Workbook()

    print("Building Setup tab...")
    build_setup_tab(wb)

    print("Building 30-Day Tracker tab...")
    build_tracker_tab(wb)

    print("Building Dashboard tab...")
    build_dashboard_tab(wb)

    print("Building Weekly Review tab...")
    build_weekly_review_tab(wb)

    print("Building Habit Ideas Bank tab...")
    build_habit_ideas_tab(wb)

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "30-Day-Habit-Tracker.xlsx")
    wb.save(output_path)
    print(f"\nWorkbook saved to: {output_path}")
    print("Done!")


if __name__ == "__main__":
    main()
