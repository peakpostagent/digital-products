"""
Build the Freelancer's Business-in-a-Box Excel workbook.
Creates a professional 7-tab workbook with formulas, data validation,
conditional formatting, and sample data.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, timedelta
import os

# ── Color Palette ──
NAVY = "1A365D"
BLUE = "2B6CB0"
LIGHT_BG = "F7FAFC"
WHITE = "FFFFFF"
LIGHT_BLUE = "EBF8FF"
ACCENT_GREEN = "38A169"
ACCENT_RED = "E53E3E"
ACCENT_YELLOW = "D69E2E"
ACCENT_ORANGE = "DD6B20"
GRAY = "A0AEC0"
LIGHT_GRAY = "E2E8F0"
DARK_TEXT = "1A202C"
MED_GRAY = "718096"

# ── Reusable Styles ──
header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=NAVY, size=11)
subheader_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
body_font = Font(name="Calibri", color=DARK_TEXT, size=10)
currency_fmt = '"$"#,##0.00'
pct_fmt = "0.0%"
date_fmt = "MM/DD/YYYY"
thin_border = Border(
    left=Side(style="thin", color=LIGHT_GRAY),
    right=Side(style="thin", color=LIGHT_GRAY),
    top=Side(style="thin", color=LIGHT_GRAY),
    bottom=Side(style="thin", color=LIGHT_GRAY),
)
blue_accent_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
green_fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
red_fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
yellow_fill = PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid")
light_row_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
formula_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_row(ws, row, max_col, alt=False):
    """Apply data row styling."""
    fill = light_row_fill if alt else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_metric_box(ws, row, col, label, value_cell_or_formula, is_formula=True):
    """Create a styled metric box (label + value)."""
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
    label_cell.fill = blue_accent_fill
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = thin_border

    val_cell = ws.cell(row=row + 1, column=col)
    if is_formula:
        val_cell.value = value_cell_or_formula
    else:
        val_cell.value = value_cell_or_formula
    val_cell.font = Font(name="Calibri", bold=True, color=NAVY, size=16)
    val_cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.border = thin_border
    val_cell.number_format = currency_fmt


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ═══════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────────
# TAB 2: Client CRM (create first so we can reference it)
# ───────────────────────────────────────────────────────────
ws_crm = wb.active
ws_crm.title = "Client CRM"
ws_crm.sheet_properties.tabColor = BLUE

crm_headers = [
    "Client Name", "Company", "Email", "Phone", "Industry",
    "Source", "Status", "First Contact", "Last Contact",
    "Lifetime Value", "Notes", "Next Follow-up"
]
for col, h in enumerate(crm_headers, 1):
    ws_crm.cell(row=1, column=col, value=h)
style_header_row(ws_crm, 1, len(crm_headers))

# Data validation for Source and Status
dv_source = DataValidation(type="list", formula1='"Referral,Website,Social Media,Cold Outreach,Networking,Conference,Other"', allow_blank=True)
dv_source.prompt = "Select lead source"
dv_source.promptTitle = "Source"
ws_crm.add_data_validation(dv_source)
dv_source.add("F2:F200")

dv_status = DataValidation(type="list", formula1='"Active,Inactive,Lead,Past"', allow_blank=True)
dv_status.prompt = "Select client status"
dv_status.promptTitle = "Status"
ws_crm.add_data_validation(dv_status)
dv_status.add("G2:G200")

# Sample CRM data
crm_data = [
    ["Sarah Mitchell", "BrightPath Marketing", "sarah@brightpath.co", "(555) 234-8901", "Marketing",
     "Referral", "Active", "2025-03-15", "2026-02-28",
     None, "Great communicator, prefers Slack", "2026-03-20"],
    ["David Chen", "Chen Legal Group", "david@chenlegal.com", "(555) 456-7890", "Legal",
     "Website", "Active", "2025-06-01", "2026-03-10",
     None, "Needs quick turnaround, pays on time", "2026-04-01"],
    ["Maria Rodriguez", "Bloom & Grow Cafe", "maria@bloomgrow.com", "(555) 789-0123", "Food & Bev",
     "Social Media", "Lead", "2026-02-14", "2026-03-01",
     None, "Interested in brand redesign, follow up Q2", "2026-03-25"],
]

for i, row_data in enumerate(crm_data, 2):
    for j, val in enumerate(row_data, 1):
        cell = ws_crm.cell(row=i, column=j, value=val)
    # Lifetime Value formula: sum of project values from Project Pipeline
    ws_crm.cell(row=i, column=10).value = f'=SUMIFS(\'Project Pipeline\'!I:I,\'Project Pipeline\'!B:B,A{i})'
    ws_crm.cell(row=i, column=10).number_format = currency_fmt
    ws_crm.cell(row=i, column=10).fill = formula_fill

    # Format dates
    for date_col in [8, 9, 12]:
        ws_crm.cell(row=i, column=date_col).number_format = date_fmt

    style_data_row(ws_crm, i, len(crm_headers), alt=(i % 2 == 0))

# Conditional formatting: overdue follow-ups (past today) in red
ws_crm.conditional_formatting.add(
    "L2:L200",
    FormulaRule(formula=["AND(L2<>\"\",$L2<TODAY())"],
                fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
                font=Font(color=ACCENT_RED))
)

# Status color coding
ws_crm.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Active"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")))
ws_crm.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Lead"'], fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")))
ws_crm.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Inactive"'], fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")))
ws_crm.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Past"'], fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")))

set_col_widths(ws_crm, {"A": 18, "B": 22, "C": 25, "D": 16, "E": 14, "F": 15, "G": 12, "H": 14, "I": 14, "J": 16, "K": 30, "L": 16})
ws_crm.freeze_panes = "A2"
ws_crm.auto_filter.ref = f"A1:L{len(crm_data)+1}"

# ───────────────────────────────────────────────────────────
# TAB 3: Project Pipeline
# ───────────────────────────────────────────────────────────
ws_proj = wb.create_sheet("Project Pipeline")
ws_proj.sheet_properties.tabColor = ACCENT_GREEN

proj_headers = [
    "Project Name", "Client", "Status", "Start Date", "Deadline",
    "Est. Hours", "Actual Hours", "Hourly Rate", "Project Value",
    "Amount Paid", "Amount Due", "Priority", "Notes"
]
for col, h in enumerate(proj_headers, 1):
    ws_proj.cell(row=1, column=col, value=h)
style_header_row(ws_proj, 1, len(proj_headers))

# Data validations
dv_proj_status = DataValidation(type="list", formula1='"Lead,Proposal,Active,Review,Complete,Cancelled"', allow_blank=True)
ws_proj.add_data_validation(dv_proj_status)
dv_proj_status.add("C2:C200")

dv_priority = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
ws_proj.add_data_validation(dv_priority)
dv_priority.add("L2:L200")

# Client dropdown from CRM
dv_client = DataValidation(type="list", formula1="='Client CRM'!$A$2:$A$200", allow_blank=True)
ws_proj.add_data_validation(dv_client)
dv_client.add("B2:B200")

proj_data = [
    ["Website Redesign", "Sarah Mitchell", "Active", "2026-01-15", "2026-04-15",
     80, 45, 95, None, 3800, None, "High", "Phase 2 starting next week"],
    ["Brand Identity Package", "Maria Rodriguez", "Proposal", "2026-04-01", "2026-05-30",
     40, 0, 85, None, 0, None, "Medium", "Waiting on client approval"],
    ["Legal Blog Content", "David Chen", "Active", "2026-02-01", "2026-06-30",
     120, 68, 75, None, 4500, None, "Medium", "Monthly retainer - 10 posts/mo"],
    ["Email Campaign Setup", "Sarah Mitchell", "Complete", "2025-10-01", "2025-12-15",
     30, 28, 95, None, 2660, None, "Low", "Delivered on time, client happy"],
]

for i, row_data in enumerate(proj_data, 2):
    for j, val in enumerate(row_data, 1):
        cell = ws_proj.cell(row=i, column=j, value=val)

    # Project Value formula = Est. Hours * Hourly Rate
    ws_proj.cell(row=i, column=9).value = f"=F{i}*H{i}"
    ws_proj.cell(row=i, column=9).number_format = currency_fmt
    ws_proj.cell(row=i, column=9).fill = formula_fill

    # Amount Due formula = Project Value - Amount Paid
    ws_proj.cell(row=i, column=11).value = f"=I{i}-J{i}"
    ws_proj.cell(row=i, column=11).number_format = currency_fmt
    ws_proj.cell(row=i, column=11).fill = formula_fill

    # Currency format
    ws_proj.cell(row=i, column=8).number_format = currency_fmt
    ws_proj.cell(row=i, column=10).number_format = currency_fmt

    # Date format
    for dc in [4, 5]:
        ws_proj.cell(row=i, column=dc).number_format = date_fmt

    style_data_row(ws_proj, i, len(proj_headers), alt=(i % 2 == 0))

# Conditional formatting for status
ws_proj.conditional_formatting.add("C2:C200", CellIsRule(operator="equal", formula=['"Active"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")))
ws_proj.conditional_formatting.add("C2:C200", CellIsRule(operator="equal", formula=['"Complete"'], fill=PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")))
ws_proj.conditional_formatting.add("C2:C200", CellIsRule(operator="equal", formula=['"Proposal"'], fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")))
ws_proj.conditional_formatting.add("C2:C200", CellIsRule(operator="equal", formula=['"Cancelled"'], fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")))
ws_proj.conditional_formatting.add("C2:C200", CellIsRule(operator="equal", formula=['"Lead"'], fill=PatternFill(start_color="E9D8FD", end_color="E9D8FD", fill_type="solid")))

# Priority coloring
ws_proj.conditional_formatting.add("L2:L200", CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"), font=Font(bold=True, color=ACCENT_RED)))
ws_proj.conditional_formatting.add("L2:L200", CellIsRule(operator="equal", formula=['"Medium"'], fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"), font=Font(bold=True, color=ACCENT_YELLOW)))
ws_proj.conditional_formatting.add("L2:L200", CellIsRule(operator="equal", formula=['"Low"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"), font=Font(bold=True, color=ACCENT_GREEN)))

# Overdue deadline formatting
ws_proj.conditional_formatting.add(
    "E2:E200",
    FormulaRule(formula=["AND(E2<>\"\",E2<TODAY(),C2<>\"Complete\",C2<>\"Cancelled\")"],
                fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
                font=Font(color=ACCENT_RED, bold=True))
)

set_col_widths(ws_proj, {"A": 22, "B": 18, "C": 14, "D": 14, "E": 14, "F": 12, "G": 12, "H": 13, "I": 15, "J": 14, "K": 14, "L": 11, "M": 30})
ws_proj.freeze_panes = "A2"
ws_proj.auto_filter.ref = f"A1:M{len(proj_data)+1}"

# ───────────────────────────────────────────────────────────
# TAB 4: Invoice Tracker
# ───────────────────────────────────────────────────────────
ws_inv = wb.create_sheet("Invoice Tracker")
ws_inv.sheet_properties.tabColor = ACCENT_YELLOW

inv_headers = [
    "Invoice #", "Client", "Project", "Date Sent", "Due Date",
    "Amount", "Status", "Payment Date", "Payment Method",
    "Days to Pay", "Late Fee", "Notes"
]
for col, h in enumerate(inv_headers, 1):
    ws_inv.cell(row=1, column=col, value=h)
style_header_row(ws_inv, 1, len(inv_headers))

dv_inv_status = DataValidation(type="list", formula1='"Draft,Sent,Paid,Overdue,Cancelled"', allow_blank=True)
ws_inv.add_data_validation(dv_inv_status)
dv_inv_status.add("G2:G200")

dv_pay_method = DataValidation(type="list", formula1='"Bank Transfer,PayPal,Stripe,Check,Venmo,Zelle,Cash,Other"', allow_blank=True)
ws_inv.add_data_validation(dv_pay_method)
dv_pay_method.add("I2:I200")

inv_data = [
    ["INV-2026-001", "Sarah Mitchell", "Website Redesign", "2026-01-20", "2026-02-19", 3800, "Paid", "2026-02-15", "Bank Transfer", None, 0, "Deposit - 50%"],
    ["INV-2026-002", "David Chen", "Legal Blog Content", "2026-02-01", "2026-03-03", 2250, "Paid", "2026-03-01", "PayPal", None, 0, "Feb retainer"],
    ["INV-2026-003", "David Chen", "Legal Blog Content", "2026-03-01", "2026-03-31", 2250, "Sent", "", "","", 0, "Mar retainer"],
    ["INV-2026-004", "Sarah Mitchell", "Website Redesign", "2026-03-15", "2026-04-14", 3800, "Sent", "", "", "", 0, "Final 50%"],
    ["INV-2026-005", "Sarah Mitchell", "Email Campaign Setup", "2025-12-18", "2026-01-17", 2660, "Paid", "2026-01-20", "Stripe", None, 0, "Final payment"],
]

for i, row_data in enumerate(inv_data, 2):
    for j, val in enumerate(row_data, 1):
        ws_inv.cell(row=i, column=j, value=val)

    ws_inv.cell(row=i, column=6).number_format = currency_fmt
    ws_inv.cell(row=i, column=11).number_format = currency_fmt
    for dc in [4, 5, 8]:
        ws_inv.cell(row=i, column=dc).number_format = date_fmt

    # Days to Pay formula
    ws_inv.cell(row=i, column=10).value = f'=IF(H{i}<>"",H{i}-D{i},IF(G{i}="Paid","",TODAY()-D{i}))'
    ws_inv.cell(row=i, column=10).fill = formula_fill

    style_data_row(ws_inv, i, len(inv_headers), alt=(i % 2 == 0))

# Status conditional formatting
ws_inv.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Paid"'], fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")))
ws_inv.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Sent"'], fill=PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")))
ws_inv.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Overdue"'], fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"), font=Font(bold=True, color=ACCENT_RED)))
ws_inv.conditional_formatting.add("G2:G200", CellIsRule(operator="equal", formula=['"Draft"'], fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")))

# Summary row
summary_row = len(inv_data) + 3
ws_inv.cell(row=summary_row, column=1, value="SUMMARY").font = Font(name="Calibri", bold=True, color=NAVY, size=12)

labels = ["Total Invoiced:", "Total Paid:", "Total Outstanding:", "Total Overdue:"]
formulas = [
    f'=SUM(F2:F{len(inv_data)+1})',
    f'=SUMIF(G2:G{len(inv_data)+1},"Paid",F2:F{len(inv_data)+1})',
    f'=SUMIF(G2:G{len(inv_data)+1},"Sent",F2:F{len(inv_data)+1})',
    f'=SUMIF(G2:G{len(inv_data)+1},"Overdue",F2:F{len(inv_data)+1})',
]
for idx, (lbl, fml) in enumerate(zip(labels, formulas)):
    r = summary_row + 1 + idx
    ws_inv.cell(row=r, column=1, value=lbl).font = subheader_font
    ws_inv.cell(row=r, column=2, value=fml).number_format = currency_fmt
    ws_inv.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=12, color=NAVY)
    ws_inv.cell(row=r, column=2).fill = formula_fill

set_col_widths(ws_inv, {"A": 16, "B": 18, "C": 22, "D": 14, "E": 14, "F": 14, "G": 12, "H": 14, "I": 16, "J": 12, "K": 10, "L": 25})
ws_inv.freeze_panes = "A2"

# ───────────────────────────────────────────────────────────
# TAB 5: Income & Expenses
# ───────────────────────────────────────────────────────────
ws_ie = wb.create_sheet("Income & Expenses")
ws_ie.sheet_properties.tabColor = ACCENT_GREEN

# ── INCOME SECTION ──
ws_ie.cell(row=1, column=1, value="INCOME").font = Font(name="Calibri", bold=True, color=WHITE, size=14)
ws_ie.cell(row=1, column=1).fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
ws_ie.merge_cells("A1:G1")

inc_headers = ["Date", "Client", "Project", "Category", "Amount", "Payment Method", "Tax Deductible"]
for col, h in enumerate(inc_headers, 1):
    ws_ie.cell(row=2, column=col, value=h)
style_header_row(ws_ie, 2, len(inc_headers))

dv_inc_cat = DataValidation(type="list", formula1='"Freelance,Retainer,Passive,Consultation"', allow_blank=True)
ws_ie.add_data_validation(dv_inc_cat)
dv_inc_cat.add("D3:D200")

dv_tax_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_ie.add_data_validation(dv_tax_yn)
dv_tax_yn.add("G3:G200")

income_data = [
    ["2026-01-15", "Sarah Mitchell", "Website Redesign", "Freelance", 3800, "Bank Transfer", "No"],
    ["2026-01-20", "David Chen", "Legal Blog Content", "Retainer", 2250, "PayPal", "No"],
    ["2026-02-01", "", "Online Course Sales", "Passive", 340, "Stripe", "No"],
    ["2026-02-15", "David Chen", "Legal Blog Content", "Retainer", 2250, "PayPal", "No"],
    ["2026-02-20", "Sarah Mitchell", "Email Campaign Setup", "Freelance", 2660, "Stripe", "No"],
    ["2026-03-01", "", "Online Course Sales", "Passive", 415, "Stripe", "No"],
    ["2026-03-05", "David Chen", "Legal Blog Content", "Retainer", 2250, "PayPal", "No"],
    ["2026-03-10", "Maria Rodriguez", "Brand Consultation", "Consultation", 250, "Venmo", "No"],
]

for i, row_data in enumerate(income_data, 3):
    for j, val in enumerate(row_data, 1):
        ws_ie.cell(row=i, column=j, value=val)
    ws_ie.cell(row=i, column=1).number_format = date_fmt
    ws_ie.cell(row=i, column=5).number_format = currency_fmt
    style_data_row(ws_ie, i, len(inc_headers), alt=(i % 2 == 0))

inc_summary_row = 3 + len(income_data) + 1
ws_ie.cell(row=inc_summary_row, column=4, value="Total Income:").font = Font(name="Calibri", bold=True, color=NAVY, size=11)
ws_ie.cell(row=inc_summary_row, column=5).value = f"=SUM(E3:E{2+len(income_data)})"
ws_ie.cell(row=inc_summary_row, column=5).number_format = currency_fmt
ws_ie.cell(row=inc_summary_row, column=5).font = Font(name="Calibri", bold=True, size=12, color=ACCENT_GREEN)
ws_ie.cell(row=inc_summary_row, column=5).fill = formula_fill

# ── EXPENSE SECTION ──
exp_start = inc_summary_row + 2
ws_ie.cell(row=exp_start, column=1, value="EXPENSES").font = Font(name="Calibri", bold=True, color=WHITE, size=14)
ws_ie.cell(row=exp_start, column=1).fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
ws_ie.merge_cells(f"A{exp_start}:H{exp_start}")

exp_headers = ["Date", "Vendor", "Category", "Amount", "Tax Deductible", "Receipt", "Notes", ""]
exp_header_row = exp_start + 1
for col, h in enumerate(exp_headers, 1):
    ws_ie.cell(row=exp_header_row, column=col, value=h)
style_header_row(ws_ie, exp_header_row, 7)

dv_exp_cat = DataValidation(type="list", formula1='"Software,Hardware,Office,Marketing,Travel,Education,Insurance,Professional Services,Other"', allow_blank=True)
ws_ie.add_data_validation(dv_exp_cat)
dv_exp_cat.add(f"C{exp_header_row+1}:C{exp_header_row+200}")

dv_tax_exp = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_ie.add_data_validation(dv_tax_exp)
dv_tax_exp.add(f"E{exp_header_row+1}:E{exp_header_row+200}")

dv_receipt = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_ie.add_data_validation(dv_receipt)
dv_receipt.add(f"F{exp_header_row+1}:F{exp_header_row+200}")

expense_data = [
    ["2026-01-05", "Adobe", "Software", 54.99, "Yes", "Yes", "Creative Cloud monthly"],
    ["2026-01-05", "Notion", "Software", 10, "Yes", "Yes", "Workspace Pro"],
    ["2026-01-10", "WeWork", "Office", 350, "Yes", "Yes", "Hot desk membership"],
    ["2026-01-15", "Google Ads", "Marketing", 150, "Yes", "Yes", "Portfolio promotion"],
    ["2026-02-01", "Udemy", "Education", 89.99, "Yes", "Yes", "Advanced UX course"],
    ["2026-02-05", "Adobe", "Software", 54.99, "Yes", "Yes", "Creative Cloud monthly"],
    ["2026-02-10", "WeWork", "Office", 350, "Yes", "Yes", "Hot desk membership"],
    ["2026-02-20", "Delta Airlines", "Travel", 380, "Yes", "Yes", "Client meeting - NYC"],
    ["2026-03-01", "Professional Liability Ins.", "Insurance", 125, "Yes", "Yes", "Monthly premium"],
    ["2026-03-05", "Adobe", "Software", 54.99, "Yes", "Yes", "Creative Cloud monthly"],
]

for i, row_data in enumerate(expense_data):
    r = exp_header_row + 1 + i
    for j, val in enumerate(row_data, 1):
        ws_ie.cell(row=r, column=j, value=val)
    ws_ie.cell(row=r, column=1).number_format = date_fmt
    ws_ie.cell(row=r, column=4).number_format = currency_fmt
    style_data_row(ws_ie, r, 7, alt=(i % 2 == 0))

exp_summary_row = exp_header_row + len(expense_data) + 1
exp_data_end = exp_header_row + len(expense_data)
ws_ie.cell(row=exp_summary_row, column=3, value="Total Expenses:").font = Font(name="Calibri", bold=True, color=NAVY, size=11)
ws_ie.cell(row=exp_summary_row, column=4).value = f"=SUM(D{exp_header_row+1}:D{exp_data_end})"
ws_ie.cell(row=exp_summary_row, column=4).number_format = currency_fmt
ws_ie.cell(row=exp_summary_row, column=4).font = Font(name="Calibri", bold=True, size=12, color=ACCENT_RED)
ws_ie.cell(row=exp_summary_row, column=4).fill = formula_fill

ws_ie.cell(row=exp_summary_row+1, column=3, value="Tax Deductible:").font = Font(name="Calibri", bold=True, color=NAVY, size=11)
ws_ie.cell(row=exp_summary_row+1, column=4).value = f'=SUMIF(E{exp_header_row+1}:E{exp_data_end},"Yes",D{exp_header_row+1}:D{exp_data_end})'
ws_ie.cell(row=exp_summary_row+1, column=4).number_format = currency_fmt
ws_ie.cell(row=exp_summary_row+1, column=4).fill = formula_fill

# Profit/Loss
pl_row = exp_summary_row + 3
ws_ie.cell(row=pl_row, column=1, value="PROFIT / LOSS").font = Font(name="Calibri", bold=True, color=WHITE, size=14)
ws_ie.cell(row=pl_row, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws_ie.merge_cells(f"A{pl_row}:D{pl_row}")

ws_ie.cell(row=pl_row+1, column=1, value="Total Income:").font = subheader_font
ws_ie.cell(row=pl_row+1, column=2).value = f"=E{inc_summary_row}"
ws_ie.cell(row=pl_row+1, column=2).number_format = currency_fmt
ws_ie.cell(row=pl_row+1, column=2).fill = formula_fill

ws_ie.cell(row=pl_row+2, column=1, value="Total Expenses:").font = subheader_font
ws_ie.cell(row=pl_row+2, column=2).value = f"=D{exp_summary_row}"
ws_ie.cell(row=pl_row+2, column=2).number_format = currency_fmt
ws_ie.cell(row=pl_row+2, column=2).fill = formula_fill

ws_ie.cell(row=pl_row+3, column=1, value="Net Profit:").font = Font(name="Calibri", bold=True, color=NAVY, size=14)
ws_ie.cell(row=pl_row+3, column=2).value = f"=B{pl_row+1}-B{pl_row+2}"
ws_ie.cell(row=pl_row+3, column=2).number_format = currency_fmt
ws_ie.cell(row=pl_row+3, column=2).font = Font(name="Calibri", bold=True, size=14, color=ACCENT_GREEN)
ws_ie.cell(row=pl_row+3, column=2).fill = formula_fill

set_col_widths(ws_ie, {"A": 14, "B": 24, "C": 18, "D": 14, "E": 16, "F": 10, "G": 28, "H": 5})
ws_ie.freeze_panes = "A3"

# ───────────────────────────────────────────────────────────
# TAB 6: Rate Calculator
# ───────────────────────────────────────────────────────────
ws_rate = wb.create_sheet("Rate Calculator")
ws_rate.sheet_properties.tabColor = "805AD5"

# Title
ws_rate.merge_cells("A1:F1")
ws_rate.cell(row=1, column=1, value="FREELANCE RATE CALCULATOR").font = Font(name="Calibri", bold=True, color=WHITE, size=16)
ws_rate.cell(row=1, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws_rate.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws_rate.row_dimensions[1].height = 40

# Input Section
ws_rate.merge_cells("A3:C3")
ws_rate.cell(row=3, column=1, value="YOUR INPUTS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_rate.cell(row=3, column=1).fill = blue_accent_fill

inputs = [
    ("Desired Annual Income", 75000, currency_fmt),
    ("Working Weeks / Year", 48, "0"),
    ("Billable Hours / Week", 30, "0"),
    ("Annual Business Expenses", 12000, currency_fmt),
    ("Tax Rate", 0.25, pct_fmt),
    ("Desired Profit Margin", 0.15, pct_fmt),
]

for i, (label, default, fmt) in enumerate(inputs):
    r = 4 + i
    ws_rate.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws_rate.cell(row=r, column=1).alignment = Alignment(indent=1)
    cell = ws_rate.cell(row=r, column=3, value=default)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=12, color=BLUE)
    cell.fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
    cell.border = Border(
        left=Side(style="medium", color=BLUE),
        right=Side(style="medium", color=BLUE),
        top=Side(style="medium", color=BLUE),
        bottom=Side(style="medium", color=BLUE),
    )

# Calculated Outputs
ws_rate.merge_cells("A12:C12")
ws_rate.cell(row=12, column=1, value="CALCULATED RATES").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_rate.cell(row=12, column=1).fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")

outputs = [
    ("Required Revenue (pre-tax)", "=(C4+C7)/(1-C8)", currency_fmt),
    ("Billable Hours / Year", "=C5*C6", "0"),
    ("Minimum Hourly Rate", "=C13/C14", currency_fmt),
    ("Suggested Hourly Rate (with margin)", "=C15*(1+C9)", currency_fmt),
    ("Day Rate (8 hrs)", "=C16*8", currency_fmt),
    ("Weekly Rate", "=C16*C6", currency_fmt),
    ("Monthly Retainer Rate", "=C16*C6*4.33", currency_fmt),
]

for i, (label, formula, fmt) in enumerate(outputs):
    r = 13 + i
    ws_rate.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws_rate.cell(row=r, column=1).alignment = Alignment(indent=1)
    cell = ws_rate.cell(row=r, column=3, value=formula)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=12, color=ACCENT_GREEN)
    cell.fill = formula_fill
    cell.border = thin_border

# Utilization Comparison Table
ws_rate.merge_cells("A22:F22")
ws_rate.cell(row=22, column=1, value="RATE COMPARISON BY UTILIZATION").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_rate.cell(row=22, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

comp_headers = ["Utilization Rate", "Billable Hours/Year", "Hourly Rate", "Day Rate", "Weekly Rate", "Monthly Retainer"]
for col, h in enumerate(comp_headers, 1):
    ws_rate.cell(row=23, column=col, value=h)
style_header_row(ws_rate, 23, len(comp_headers))

utilizations = [0.6, 0.7, 0.8, 0.9]
for i, util in enumerate(utilizations):
    r = 24 + i
    ws_rate.cell(row=r, column=1, value=util).number_format = pct_fmt
    ws_rate.cell(row=r, column=2).value = f"=C5*C6*A{r}"
    ws_rate.cell(row=r, column=2).number_format = "0"
    ws_rate.cell(row=r, column=3).value = f"=C13/B{r}"
    ws_rate.cell(row=r, column=3).number_format = currency_fmt
    ws_rate.cell(row=r, column=4).value = f"=C{r}*8"
    ws_rate.cell(row=r, column=4).number_format = currency_fmt
    ws_rate.cell(row=r, column=5).value = f"=C{r}*C6"
    ws_rate.cell(row=r, column=5).number_format = currency_fmt
    ws_rate.cell(row=r, column=6).value = f"=C{r}*C6*4.33"
    ws_rate.cell(row=r, column=6).number_format = currency_fmt
    style_data_row(ws_rate, r, 6, alt=(i % 2 == 0))
    for c in range(2, 7):
        ws_rate.cell(row=r, column=c).fill = formula_fill

# What-If Section
ws_rate.merge_cells("A30:F30")
ws_rate.cell(row=30, column=1, value="WHAT-IF SCENARIOS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_rate.cell(row=30, column=1).fill = PatternFill(start_color="805AD5", end_color="805AD5", fill_type="solid")

ws_rate.cell(row=31, column=1, value="If I raise my rate by:").font = subheader_font
ws_rate.cell(row=31, column=3, value=10).number_format = currency_fmt
ws_rate.cell(row=31, column=3).fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
ws_rate.cell(row=31, column=3).border = Border(left=Side(style="medium", color="805AD5"), right=Side(style="medium", color="805AD5"), top=Side(style="medium", color="805AD5"), bottom=Side(style="medium", color="805AD5"))

ws_rate.cell(row=32, column=1, value="New hourly rate:").font = subheader_font
ws_rate.cell(row=32, column=3).value = "=C16+C31"
ws_rate.cell(row=32, column=3).number_format = currency_fmt
ws_rate.cell(row=32, column=3).fill = formula_fill

ws_rate.cell(row=33, column=1, value="Additional annual income:").font = subheader_font
ws_rate.cell(row=33, column=3).value = "=C31*C14"
ws_rate.cell(row=33, column=3).number_format = currency_fmt
ws_rate.cell(row=33, column=3).fill = formula_fill

ws_rate.cell(row=34, column=1, value="New annual revenue:").font = subheader_font
ws_rate.cell(row=34, column=3).value = "=C32*C14"
ws_rate.cell(row=34, column=3).number_format = currency_fmt
ws_rate.cell(row=34, column=3).fill = formula_fill

set_col_widths(ws_rate, {"A": 30, "B": 20, "C": 20, "D": 16, "E": 16, "F": 18})

# ───────────────────────────────────────────────────────────
# TAB 7: Proposal Template
# ───────────────────────────────────────────────────────────
ws_prop = wb.create_sheet("Proposal Template")
ws_prop.sheet_properties.tabColor = NAVY

# Set print area and page setup
ws_prop.page_setup.paperSize = ws_prop.PAPERSIZE_LETTER
ws_prop.page_setup.orientation = "portrait"
ws_prop.page_margins.left = 0.75
ws_prop.page_margins.right = 0.75
ws_prop.page_margins.top = 0.5
ws_prop.page_margins.bottom = 0.5

# Company Header
ws_prop.merge_cells("A1:F1")
ws_prop.cell(row=1, column=1, value="[YOUR COMPANY NAME]").font = Font(name="Calibri", bold=True, color=WHITE, size=22)
ws_prop.cell(row=1, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws_prop.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws_prop.row_dimensions[1].height = 50

ws_prop.merge_cells("A2:F2")
ws_prop.cell(row=2, column=1, value="[your@email.com] | [yourwebsite.com] | [(555) 000-0000]").font = Font(name="Calibri", color=BLUE, size=10)
ws_prop.cell(row=2, column=1).alignment = Alignment(horizontal="center")

# Proposal Title
ws_prop.merge_cells("A4:F4")
ws_prop.cell(row=4, column=1, value="PROJECT PROPOSAL").font = Font(name="Calibri", bold=True, color=NAVY, size=18)
ws_prop.cell(row=4, column=1).alignment = Alignment(horizontal="center")

# Client Info
ws_prop.merge_cells("A6:C6")
ws_prop.cell(row=6, column=1, value="PREPARED FOR:").font = Font(name="Calibri", bold=True, color=WHITE, size=11)
ws_prop.cell(row=6, column=1).fill = blue_accent_fill

ws_prop.cell(row=7, column=1, value="Client Name:").font = subheader_font
ws_prop.cell(row=7, column=2, value="Sarah Mitchell").font = body_font
ws_prop.cell(row=8, column=1, value="Company:").font = subheader_font
ws_prop.cell(row=8, column=2, value="BrightPath Marketing").font = body_font
ws_prop.cell(row=9, column=1, value="Date:").font = subheader_font
ws_prop.cell(row=9, column=2, value="March 17, 2026").font = body_font
ws_prop.cell(row=10, column=1, value="Valid Until:").font = subheader_font
ws_prop.cell(row=10, column=2, value="April 17, 2026").font = body_font

# Project Info
ws_prop.merge_cells("D6:F6")
ws_prop.cell(row=6, column=4, value="PROJECT DETAILS:").font = Font(name="Calibri", bold=True, color=WHITE, size=11)
ws_prop.cell(row=6, column=4).fill = blue_accent_fill

ws_prop.cell(row=7, column=4, value="Project:").font = subheader_font
ws_prop.cell(row=7, column=5, value="Website Redesign").font = body_font
ws_prop.cell(row=8, column=4, value="Timeline:").font = subheader_font
ws_prop.cell(row=8, column=5, value="8-10 weeks").font = body_font
ws_prop.cell(row=9, column=4, value="Proposal #:").font = subheader_font
ws_prop.cell(row=9, column=5, value="PROP-2026-007").font = body_font

# Project Overview
ws_prop.merge_cells("A12:F12")
ws_prop.cell(row=12, column=1, value="PROJECT OVERVIEW").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_prop.cell(row=12, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

ws_prop.merge_cells("A13:F15")
ws_prop.cell(row=13, column=1, value="A complete redesign of the BrightPath Marketing website to improve user experience, modernize the visual design, and optimize for lead generation. The project includes responsive design, SEO optimization, and integration with existing marketing tools.").font = body_font
ws_prop.cell(row=13, column=1).alignment = Alignment(wrap_text=True, vertical="top")

# Scope of Work Table
ws_prop.merge_cells("A17:F17")
ws_prop.cell(row=17, column=1, value="SCOPE OF WORK").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_prop.cell(row=17, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

scope_headers = ["#", "Deliverable", "Description", "Timeline", "Price", ""]
for col, h in enumerate(scope_headers, 1):
    ws_prop.cell(row=18, column=col, value=h)
style_header_row(ws_prop, 18, 5)

scope_items = [
    [1, "Discovery & Research", "Stakeholder interviews, competitor analysis, user research", "Week 1-2", 1200],
    [2, "Wireframes & Prototypes", "Low-fi wireframes, interactive prototype, client review", "Week 3-4", 1800],
    [3, "Visual Design", "High-fidelity mockups for all pages, style guide", "Week 5-6", 2400],
    [4, "Development", "Responsive HTML/CSS/JS build, CMS integration", "Week 7-9", 3200],
    [5, "Testing & Launch", "QA testing, bug fixes, deployment, training", "Week 10", 1000],
]

for i, item in enumerate(scope_items):
    r = 19 + i
    for j, val in enumerate(item, 1):
        ws_prop.cell(row=r, column=j, value=val)
    ws_prop.cell(row=r, column=5).number_format = currency_fmt
    style_data_row(ws_prop, r, 5, alt=(i % 2 == 0))

total_row = 19 + len(scope_items)
ws_prop.cell(row=total_row, column=4, value="TOTAL:").font = Font(name="Calibri", bold=True, color=NAVY, size=12)
ws_prop.cell(row=total_row, column=5).value = f"=SUM(E19:E{total_row-1})"
ws_prop.cell(row=total_row, column=5).number_format = currency_fmt
ws_prop.cell(row=total_row, column=5).font = Font(name="Calibri", bold=True, color=NAVY, size=14)
ws_prop.cell(row=total_row, column=5).fill = formula_fill

# Payment Terms
pt_row = total_row + 2
ws_prop.merge_cells(f"A{pt_row}:F{pt_row}")
ws_prop.cell(row=pt_row, column=1, value="PAYMENT TERMS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_prop.cell(row=pt_row, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

terms = [
    "50% deposit required before work begins",
    "25% due upon design approval (after Phase 3)",
    "25% due upon project completion and delivery",
    "Payment due within 15 days of invoice date",
    "Late payments subject to 1.5% monthly interest",
]
for i, term in enumerate(terms):
    r = pt_row + 1 + i
    ws_prop.cell(row=r, column=1, value=f"  {i+1}.")
    ws_prop.merge_cells(f"B{r}:F{r}")
    ws_prop.cell(row=r, column=2, value=term).font = body_font

# Terms & Conditions
tc_row = pt_row + 1 + len(terms) + 1
ws_prop.merge_cells(f"A{tc_row}:F{tc_row}")
ws_prop.cell(row=tc_row, column=1, value="TERMS & CONDITIONS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_prop.cell(row=tc_row, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

conditions = [
    "This proposal is valid for 30 days from the date above.",
    "Project timeline begins upon receipt of signed agreement and deposit.",
    "Up to 2 rounds of revisions are included per deliverable phase.",
    "Additional revisions will be billed at the agreed hourly rate.",
    "Client is responsible for providing all required content and assets on time.",
    "Delays in client feedback may extend the project timeline.",
]
for i, cond in enumerate(conditions):
    r = tc_row + 1 + i
    ws_prop.cell(row=r, column=1, value=f"  {i+1}.")
    ws_prop.merge_cells(f"B{r}:F{r}")
    ws_prop.cell(row=r, column=2, value=cond).font = Font(name="Calibri", color=MED_GRAY, size=9)

# Signature
sig_row = tc_row + 1 + len(conditions) + 2
ws_prop.merge_cells(f"A{sig_row}:C{sig_row}")
ws_prop.cell(row=sig_row, column=1, value="ACCEPTED BY:").font = subheader_font
ws_prop.cell(row=sig_row+1, column=1, value="Signature: ___________________________").font = body_font
ws_prop.cell(row=sig_row+2, column=1, value="Date: ___________________________").font = body_font

ws_prop.merge_cells(f"D{sig_row}:F{sig_row}")
ws_prop.cell(row=sig_row, column=4, value="PREPARED BY:").font = subheader_font
ws_prop.cell(row=sig_row+1, column=4, value="Signature: ___________________________").font = body_font
ws_prop.cell(row=sig_row+2, column=4, value="Date: ___________________________").font = body_font

set_col_widths(ws_prop, {"A": 6, "B": 22, "C": 40, "D": 14, "E": 14, "F": 14})
ws_prop.print_area = f"A1:F{sig_row+3}"

# ───────────────────────────────────────────────────────────
# TAB 1: Dashboard (create last so formulas reference other tabs)
# ───────────────────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_properties.tabColor = NAVY

# Title
ws_dash.merge_cells("A1:H1")
ws_dash.cell(row=1, column=1, value="FREELANCER BUSINESS DASHBOARD").font = Font(name="Calibri", bold=True, color=WHITE, size=18)
ws_dash.cell(row=1, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
ws_dash.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 50

ws_dash.merge_cells("A2:H2")
ws_dash.cell(row=2, column=1, value="Your business at a glance  |  Data updates automatically from other tabs").font = Font(name="Calibri", italic=True, color=MED_GRAY, size=10)
ws_dash.cell(row=2, column=1).alignment = Alignment(horizontal="center")

# ── Key Metrics Row ──
ws_dash.merge_cells("A4:H4")
ws_dash.cell(row=4, column=1, value="KEY METRICS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_dash.cell(row=4, column=1).fill = blue_accent_fill

# Metric boxes (row 5 = label, row 6 = value)
metrics = [
    ("A", "B", "Total Revenue (YTD)", f"='Income & Expenses'!E{inc_summary_row}"),
    ("C", "D", "Outstanding Invoices", "=SUMIF('Invoice Tracker'!G:G,\"Sent\",'Invoice Tracker'!F:F)"),
    ("E", "F", "Active Projects", "=COUNTIF('Project Pipeline'!C:C,\"Active\")"),
    ("G", "H", "Monthly Avg Income", f"='Income & Expenses'!E{inc_summary_row}/MONTH(TODAY())"),
]

for (col1, col2, label, formula) in metrics:
    c1 = openpyxl.utils.column_index_from_string(col1)
    c2 = openpyxl.utils.column_index_from_string(col2)
    ws_dash.merge_cells(f"{col1}5:{col2}5")
    ws_dash.cell(row=5, column=c1, value=label)
    ws_dash.cell(row=5, column=c1).font = Font(name="Calibri", bold=True, color=WHITE, size=9)
    ws_dash.cell(row=5, column=c1).fill = blue_accent_fill
    ws_dash.cell(row=5, column=c1).alignment = Alignment(horizontal="center", vertical="center")

    ws_dash.merge_cells(f"{col1}6:{col2}6")
    ws_dash.cell(row=6, column=c1, value=formula)
    ws_dash.cell(row=6, column=c1).font = Font(name="Calibri", bold=True, color=NAVY, size=18)
    ws_dash.cell(row=6, column=c1).fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    ws_dash.cell(row=6, column=c1).alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.cell(row=6, column=c1).number_format = currency_fmt
    ws_dash.cell(row=6, column=c1).border = thin_border
    ws_dash.row_dimensions[6].height = 40

# ── Quick Stats ──
ws_dash.merge_cells("A8:D8")
ws_dash.cell(row=8, column=1, value="QUICK STATS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_dash.cell(row=8, column=1).fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")

quick_stats = [
    ("Total Clients:", "=COUNTA('Client CRM'!A2:A200)"),
    ("Active Clients:", "=COUNTIF('Client CRM'!G:G,\"Active\")"),
    ("Avg Project Value:", "=AVERAGE('Project Pipeline'!I2:I200)"),
    ("Total Expenses (YTD):", f"='Income & Expenses'!D{exp_summary_row}"),
    ("Net Profit (YTD):", f"='Income & Expenses'!B{pl_row+3}"),
    ("Profit Margin:", f"=IF('Income & Expenses'!B{pl_row+1}>0,'Income & Expenses'!B{pl_row+3}/'Income & Expenses'!B{pl_row+1},0)"),
]

for i, (label, formula) in enumerate(quick_stats):
    r = 9 + i
    ws_dash.cell(row=r, column=1, value=label).font = subheader_font
    ws_dash.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_dash.merge_cells(f"B{r}:D{r}")
    cell = ws_dash.cell(row=r, column=2, value=formula)
    cell.fill = formula_fill
    cell.border = thin_border
    if "Margin" in label:
        cell.number_format = pct_fmt
    else:
        cell.number_format = currency_fmt
    cell.font = Font(name="Calibri", bold=True, size=12, color=NAVY)

# ── Monthly Revenue Breakdown ──
ws_dash.merge_cells("F8:H8")
ws_dash.cell(row=8, column=6, value="MONTHLY REVENUE").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_dash.cell(row=8, column=6).fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")

months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]

ws_dash.cell(row=9, column=6, value="Month").font = Font(name="Calibri", bold=True, color=WHITE, size=10)
ws_dash.cell(row=9, column=6).fill = header_fill
ws_dash.cell(row=9, column=7, value="Income").font = Font(name="Calibri", bold=True, color=WHITE, size=10)
ws_dash.cell(row=9, column=7).fill = header_fill
ws_dash.cell(row=9, column=8, value="Expenses").font = Font(name="Calibri", bold=True, color=WHITE, size=10)
ws_dash.cell(row=9, column=8).fill = header_fill

for i, month in enumerate(months):
    r = 10 + i
    ws_dash.cell(row=r, column=6, value=month).font = body_font
    # SUMPRODUCT to match month of income dates
    ws_dash.cell(row=r, column=7).value = f"=SUMPRODUCT(('Income & Expenses'!A3:A{2+len(income_data)}<>\"\")*" \
                                           f"(MONTH('Income & Expenses'!A3:A{2+len(income_data)})={i+1})*" \
                                           f"('Income & Expenses'!E3:E{2+len(income_data)}))"
    ws_dash.cell(row=r, column=7).number_format = currency_fmt
    ws_dash.cell(row=r, column=7).fill = formula_fill
    ws_dash.cell(row=r, column=7).border = thin_border

    ws_dash.cell(row=r, column=8).value = f"=SUMPRODUCT(('Income & Expenses'!A{exp_header_row+1}:A{exp_data_end}<>\"\")*" \
                                           f"(MONTH('Income & Expenses'!A{exp_header_row+1}:A{exp_data_end})={i+1})*" \
                                           f"('Income & Expenses'!D{exp_header_row+1}:D{exp_data_end}))"
    ws_dash.cell(row=r, column=8).number_format = currency_fmt
    ws_dash.cell(row=r, column=8).fill = formula_fill
    ws_dash.cell(row=r, column=8).border = thin_border

    style_data_row(ws_dash, r, 8, alt=(i % 2 == 0))

# ── Pipeline Status Summary ──
ws_dash.merge_cells("A16:D16")
ws_dash.cell(row=16, column=1, value="PIPELINE STATUS").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_dash.cell(row=16, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

statuses = ["Lead", "Proposal", "Active", "Review", "Complete"]
for i, status in enumerate(statuses):
    r = 17 + i
    ws_dash.cell(row=r, column=1, value=status).font = subheader_font
    ws_dash.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_dash.cell(row=r, column=2).value = f'=COUNTIF(\'Project Pipeline\'!C:C,"{status}")'
    ws_dash.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws_dash.cell(row=r, column=2).fill = formula_fill
    ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.merge_cells(f"C{r}:D{r}")
    ws_dash.cell(row=r, column=3).value = f'=SUMIF(\'Project Pipeline\'!C:C,"{status}",\'Project Pipeline\'!I:I)'
    ws_dash.cell(row=r, column=3).number_format = currency_fmt
    ws_dash.cell(row=r, column=3).fill = formula_fill
    ws_dash.cell(row=r, column=3).border = thin_border

# ── Invoice Summary ──
ws_dash.merge_cells("A23:D23")
ws_dash.cell(row=23, column=1, value="INVOICE SUMMARY").font = Font(name="Calibri", bold=True, color=WHITE, size=12)
ws_dash.cell(row=23, column=1).fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

inv_statuses = [("Paid", ACCENT_GREEN), ("Sent", BLUE), ("Overdue", ACCENT_RED), ("Draft", GRAY)]
for i, (status, color) in enumerate(inv_statuses):
    r = 24 + i
    ws_dash.cell(row=r, column=1, value=status).font = Font(name="Calibri", bold=True, color=color, size=11)
    ws_dash.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_dash.cell(row=r, column=2).value = f'=COUNTIF(\'Invoice Tracker\'!G:G,"{status}")'
    ws_dash.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=14, color=color)
    ws_dash.cell(row=r, column=2).fill = formula_fill
    ws_dash.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.merge_cells(f"C{r}:D{r}")
    ws_dash.cell(row=r, column=3).value = f'=SUMIF(\'Invoice Tracker\'!G:G,"{status}",\'Invoice Tracker\'!F:F)'
    ws_dash.cell(row=r, column=3).number_format = currency_fmt
    ws_dash.cell(row=r, column=3).fill = formula_fill
    ws_dash.cell(row=r, column=3).border = thin_border

set_col_widths(ws_dash, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 3, "F": 14, "G": 14, "H": 14})
ws_dash.freeze_panes = "A3"

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
output_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(output_dir, "Freelancer-Business-in-a-Box.xlsx")
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
