"""
Build the Developer Project Estimation Spreadsheet (.xlsx).
Creates a professional 6-tab workbook with formulas, data validation,
conditional formatting, and sample data for freelance project estimation.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from datetime import datetime, timedelta
import os

# ── Color Palette ──
NAVY = "1A365D"
BLUE = "2B6CB0"
LIGHT_BG = "F7FAFC"
WHITE = "FFFFFF"
LIGHT_BLUE = "EBF8FF"
ACCENT_GREEN = "38A169"
ACCENT_RED = "E53E3E"
ACCENT_YELLOW = "D69E2E"
ACCENT_ORANGE = "DD6B20"
GRAY = "A0AEC0"
LIGHT_GRAY = "E2E8F0"
DARK_TEXT = "1A202C"
MED_GRAY = "718096"
PURPLE = "805AD5"

# ── Reusable Styles ──
header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, color=NAVY, size=11)
subheader_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
body_font = Font(name="Calibri", color=DARK_TEXT, size=10)
currency_fmt = '"$"#,##0.00'
pct_fmt = "0.0%"
date_fmt = "MM/DD/YYYY"
thin_border = Border(
    left=Side(style="thin", color=LIGHT_GRAY),
    right=Side(style="thin", color=LIGHT_GRAY),
    top=Side(style="thin", color=LIGHT_GRAY),
    bottom=Side(style="thin", color=LIGHT_GRAY),
)
blue_accent_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
green_fill = PatternFill(start_color=ACCENT_GREEN, end_color=ACCENT_GREEN, fill_type="solid")
red_fill = PatternFill(start_color=ACCENT_RED, end_color=ACCENT_RED, fill_type="solid")
yellow_fill = PatternFill(start_color=ACCENT_YELLOW, end_color=ACCENT_YELLOW, fill_type="solid")
light_row_fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
formula_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
input_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
input_border = Border(
    left=Side(style="medium", color=BLUE),
    right=Side(style="medium", color=BLUE),
    top=Side(style="medium", color=BLUE),
    bottom=Side(style="medium", color=BLUE),
)


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_row(ws, row, max_col, alt=False):
    """Apply data row styling."""
    fill = light_row_fill if alt else PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_section_header(ws, row, col_start, col_end, title, color=NAVY):
    """Create a merged section header."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=title)
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=12)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).border = thin_border


def style_metric_box(ws, row, col, label, value_or_formula, num_format=None):
    """Create a styled metric box (label above, value below)."""
    # Label cell
    label_cell = ws.cell(row=row, column=col, value=label)
    label_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
    label_cell.fill = blue_accent_fill
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = thin_border

    # Value cell
    val_cell = ws.cell(row=row + 1, column=col, value=value_or_formula)
    val_cell.font = Font(name="Calibri", bold=True, color=NAVY, size=16)
    val_cell.fill = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type="solid")
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.border = thin_border
    if num_format:
        val_cell.number_format = num_format


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ═══════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ───────────────────────────────────────────────────────────
# TAB 2: Feature Breakdown (create first so Dashboard can reference it)
# ───────────────────────────────────────────────────────────
ws_feat = wb.active
ws_feat.title = "Feature Breakdown"
ws_feat.sheet_properties.tabColor = BLUE

feat_headers = [
    "Feature Name", "Complexity", "Estimated Hours",
    "Priority", "Status", "Notes"
]
for col, h in enumerate(feat_headers, 1):
    ws_feat.cell(row=1, column=col, value=h)
style_header_row(ws_feat, 1, len(feat_headers))

# Data validation: Complexity
dv_complexity = DataValidation(
    type="list", formula1='"Low,Medium,High"', allow_blank=True
)
dv_complexity.prompt = "Low=2-4h, Med=4-8h, High=8-20h"
dv_complexity.promptTitle = "Complexity"
ws_feat.add_data_validation(dv_complexity)
dv_complexity.add("B2:B100")

# Data validation: Priority (MoSCoW)
dv_priority = DataValidation(
    type="list", formula1='"Must Have,Should Have,Could Have,Won\'t Have"', allow_blank=True
)
dv_priority.prompt = "Select priority level"
dv_priority.promptTitle = "Priority"
ws_feat.add_data_validation(dv_priority)
dv_priority.add("D2:D100")

# Data validation: Status
dv_feat_status = DataValidation(
    type="list", formula1='"Not Started,In Progress,Done,Blocked"', allow_blank=True
)
dv_feat_status.prompt = "Select feature status"
dv_feat_status.promptTitle = "Status"
ws_feat.add_data_validation(dv_feat_status)
dv_feat_status.add("E2:E100")

# Sample feature data
feat_data = [
    ["User Authentication (Login/Register)", "High", 16, "Must Have", "Done", "OAuth + email/password"],
    ["User Profile Page", "Medium", 6, "Must Have", "In Progress", "Avatar upload, edit details"],
    ["Product Listing Page", "Medium", 8, "Must Have", "Not Started", "Grid/list view, filters"],
    ["Shopping Cart", "High", 14, "Must Have", "Not Started", "Add/remove, qty update, persist"],
    ["Checkout & Payment Integration", "High", 20, "Must Have", "Not Started", "Stripe integration"],
    ["Search & Filtering", "Medium", 6, "Should Have", "Not Started", "Full-text search, category filter"],
    ["Admin Dashboard", "High", 12, "Should Have", "Not Started", "CRUD for products, orders"],
    ["Email Notifications", "Low", 4, "Should Have", "Not Started", "Order confirm, shipping update"],
    ["Product Reviews & Ratings", "Medium", 8, "Could Have", "Not Started", "Star rating, text review"],
    ["Wishlist / Favorites", "Low", 3, "Could Have", "Not Started", "Save items for later"],
    ["Social Media Sharing", "Low", 2, "Won't Have", "Not Started", "Share buttons on products"],
    ["Multi-language Support", "High", 18, "Won't Have", "Not Started", "i18n framework setup"],
]

for i, row_data in enumerate(feat_data, 2):
    for j, val in enumerate(row_data, 1):
        ws_feat.cell(row=i, column=j, value=val)
    style_data_row(ws_feat, i, len(feat_headers), alt=(i % 2 == 0))

# Conditional formatting: Status colors
ws_feat.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Done"'],
              fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
              font=Font(color=ACCENT_GREEN, bold=True))
)
ws_feat.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"In Progress"'],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
              font=Font(color=ACCENT_YELLOW, bold=True))
)
ws_feat.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Blocked"'],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(color=ACCENT_RED, bold=True))
)
ws_feat.conditional_formatting.add(
    "E2:E100",
    CellIsRule(operator="equal", formula=['"Not Started"'],
              fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"))
)

# Priority coloring
ws_feat.conditional_formatting.add(
    "D2:D100",
    CellIsRule(operator="equal", formula=['"Must Have"'],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(bold=True, color=ACCENT_RED))
)
ws_feat.conditional_formatting.add(
    "D2:D100",
    CellIsRule(operator="equal", formula=['"Should Have"'],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
              font=Font(bold=True, color=ACCENT_YELLOW))
)
ws_feat.conditional_formatting.add(
    "D2:D100",
    CellIsRule(operator="equal", formula=['"Could Have"'],
              fill=PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid"),
              font=Font(bold=True, color=BLUE))
)
ws_feat.conditional_formatting.add(
    "D2:D100",
    CellIsRule(operator="equal", formula=['"Won\'t Have"'],
              fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"),
              font=Font(color=MED_GRAY))
)

# Total hours row
total_row = len(feat_data) + 3
ws_feat.cell(row=total_row, column=1, value="TOTAL ESTIMATED HOURS").font = Font(
    name="Calibri", bold=True, color=NAVY, size=12
)
ws_feat.cell(row=total_row, column=3, value=f"=SUM(C2:C{len(feat_data)+1})")
ws_feat.cell(row=total_row, column=3).font = Font(
    name="Calibri", bold=True, color=NAVY, size=14
)
ws_feat.cell(row=total_row, column=3).fill = formula_fill
ws_feat.cell(row=total_row, column=3).border = thin_border
ws_feat.cell(row=total_row, column=3).number_format = "0.0"

# Must-have hours only
ws_feat.cell(row=total_row + 1, column=1, value="Must-Have Hours Only").font = subheader_font
ws_feat.cell(row=total_row + 1, column=3).value = f'=SUMIF(D2:D{len(feat_data)+1},"Must Have",C2:C{len(feat_data)+1})'
ws_feat.cell(row=total_row + 1, column=3).font = Font(name="Calibri", bold=True, color=BLUE, size=12)
ws_feat.cell(row=total_row + 1, column=3).fill = formula_fill
ws_feat.cell(row=total_row + 1, column=3).border = thin_border
ws_feat.cell(row=total_row + 1, column=3).number_format = "0.0"

# Completion percentage
ws_feat.cell(row=total_row + 2, column=1, value="Completion %").font = subheader_font
ws_feat.cell(row=total_row + 2, column=3).value = f'=IF(COUNTA(A2:A{len(feat_data)+1})=0,0,COUNTIF(E2:E{len(feat_data)+1},"Done")/COUNTA(A2:A{len(feat_data)+1}))'
ws_feat.cell(row=total_row + 2, column=3).font = Font(name="Calibri", bold=True, color=ACCENT_GREEN, size=12)
ws_feat.cell(row=total_row + 2, column=3).fill = formula_fill
ws_feat.cell(row=total_row + 2, column=3).border = thin_border
ws_feat.cell(row=total_row + 2, column=3).number_format = pct_fmt

set_col_widths(ws_feat, {"A": 36, "B": 14, "C": 16, "D": 16, "E": 14, "F": 35})
ws_feat.freeze_panes = "A2"
ws_feat.auto_filter.ref = f"A1:F{len(feat_data)+1}"


# ───────────────────────────────────────────────────────────
# TAB 3: Timeline Builder
# ───────────────────────────────────────────────────────────
ws_time = wb.create_sheet("Timeline Builder")
ws_time.sheet_properties.tabColor = ACCENT_GREEN

time_headers = [
    "Phase", "Start Date", "End Date", "Duration (days)",
    "Dependencies", "Owner", "Status"
]
for col, h in enumerate(time_headers, 1):
    ws_time.cell(row=1, column=col, value=h)
style_header_row(ws_time, 1, len(time_headers))

# Data validation for Status
dv_phase_status = DataValidation(
    type="list", formula1='"Not Started,In Progress,Done,Blocked"', allow_blank=True
)
ws_time.add_data_validation(dv_phase_status)
dv_phase_status.add("G2:G20")

# Pre-populated phases with sample dates
time_data = [
    ["Discovery & Requirements", "2026-04-14", None, 5, "None", "Lead Dev", "Not Started"],
    ["UI/UX Design", None, None, 8, "Discovery", "Designer", "Not Started"],
    ["Frontend Development", None, None, 15, "Design", "Frontend Dev", "Not Started"],
    ["Backend Development", None, None, 15, "Discovery", "Backend Dev", "Not Started"],
    ["Integration & Testing", None, None, 10, "Frontend, Backend", "Full Team", "Not Started"],
    ["Deployment & Launch", None, None, 3, "Testing", "DevOps", "Not Started"],
    ["Buffer", None, None, None, "All phases", "PM", "Not Started"],
]

for i, row_data in enumerate(time_data, 2):
    ws_time.cell(row=i, column=1, value=row_data[0])  # Phase
    ws_time.cell(row=i, column=5, value=row_data[4])  # Dependencies
    ws_time.cell(row=i, column=6, value=row_data[5])  # Owner
    ws_time.cell(row=i, column=7, value=row_data[6])  # Status

    if i == 2:
        # First phase: set start date directly
        ws_time.cell(row=i, column=2, value=row_data[1])
        ws_time.cell(row=i, column=2).number_format = date_fmt
        ws_time.cell(row=i, column=4, value=row_data[3])
        # End date = start + duration
        ws_time.cell(row=i, column=3).value = f"=B{i}+D{i}-1"
        ws_time.cell(row=i, column=3).number_format = date_fmt
        ws_time.cell(row=i, column=3).fill = formula_fill
    elif i == 8:
        # Buffer phase: auto-calculate 20% of total timeline
        ws_time.cell(row=i, column=4).value = f"=ROUND(SUM(D2:D{i-1})*0.2,0)"
        ws_time.cell(row=i, column=4).fill = formula_fill
        # Start = day after previous phase ends
        ws_time.cell(row=i, column=2).value = f"=C{i-1}+1"
        ws_time.cell(row=i, column=2).number_format = date_fmt
        ws_time.cell(row=i, column=2).fill = formula_fill
        ws_time.cell(row=i, column=3).value = f"=B{i}+D{i}-1"
        ws_time.cell(row=i, column=3).number_format = date_fmt
        ws_time.cell(row=i, column=3).fill = formula_fill
    else:
        ws_time.cell(row=i, column=4, value=row_data[3])
        # Start = day after previous phase ends
        ws_time.cell(row=i, column=2).value = f"=C{i-1}+1"
        ws_time.cell(row=i, column=2).number_format = date_fmt
        ws_time.cell(row=i, column=2).fill = formula_fill
        # End date = start + duration - 1
        ws_time.cell(row=i, column=3).value = f"=B{i}+D{i}-1"
        ws_time.cell(row=i, column=3).number_format = date_fmt
        ws_time.cell(row=i, column=3).fill = formula_fill

    style_data_row(ws_time, i, len(time_headers), alt=(i % 2 == 0))

# Status conditional formatting
ws_time.conditional_formatting.add(
    "G2:G20",
    CellIsRule(operator="equal", formula=['"Done"'],
              fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
              font=Font(color=ACCENT_GREEN, bold=True))
)
ws_time.conditional_formatting.add(
    "G2:G20",
    CellIsRule(operator="equal", formula=['"In Progress"'],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
              font=Font(color=ACCENT_YELLOW, bold=True))
)
ws_time.conditional_formatting.add(
    "G2:G20",
    CellIsRule(operator="equal", formula=['"Blocked"'],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(color=ACCENT_RED, bold=True))
)

# Gantt-chart-style visual: weeks across columns I-T
gantt_start_col = 9  # Column I
num_weeks = 12
ws_time.cell(row=1, column=gantt_start_col - 1, value="").border = thin_border

# Week headers
for w in range(num_weeks):
    col = gantt_start_col + w
    ws_time.cell(row=1, column=col, value=f"Wk {w+1}")
    ws_time.cell(row=1, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=9)
    ws_time.cell(row=1, column=col).fill = header_fill
    ws_time.cell(row=1, column=col).alignment = Alignment(horizontal="center")
    ws_time.cell(row=1, column=col).border = thin_border
    ws_time.column_dimensions[get_column_letter(col)].width = 6

# Gantt bar formulas: shade cell if week overlaps with phase dates
for i in range(2, 2 + len(time_data)):
    for w in range(num_weeks):
        col = gantt_start_col + w
        # Week start = project start + (w * 7), week end = week start + 6
        formula = (
            f'=IF(AND(B{i}<>"",D{i}<>""),'
            f'IF(AND(B$2+{w*7}<=C{i},B$2+{(w+1)*7-1}>=B{i}),"X",""),"")'
        )
        cell = ws_time.cell(row=i, column=col, value=formula)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Conditional formatting for Gantt bars (highlight cells with "X")
    gantt_range = f"{get_column_letter(gantt_start_col)}{i}:{get_column_letter(gantt_start_col + num_weeks - 1)}{i}"
    # Use different colors per phase
    colors = ["4299E1", "48BB78", "ED8936", "9F7AEA", "ED64A6", "38B2AC", "A0AEC0"]
    color_idx = (i - 2) % len(colors)
    ws_time.conditional_formatting.add(
        gantt_range,
        CellIsRule(
            operator="equal", formula=['"X"'],
            fill=PatternFill(start_color=colors[color_idx], end_color=colors[color_idx], fill_type="solid"),
            font=Font(color=colors[color_idx], size=1)
        )
    )

# Summary row
summary_row = len(time_data) + 3
ws_time.cell(row=summary_row, column=1, value="TOTAL PROJECT DURATION").font = Font(
    name="Calibri", bold=True, color=NAVY, size=12
)
ws_time.cell(row=summary_row, column=4).value = f"=SUM(D2:D{len(time_data)+1})"
ws_time.cell(row=summary_row, column=4).font = Font(name="Calibri", bold=True, color=NAVY, size=14)
ws_time.cell(row=summary_row, column=4).fill = formula_fill
ws_time.cell(row=summary_row, column=4).border = thin_border
ws_time.cell(row=summary_row, column=4).number_format = "0"

ws_time.cell(row=summary_row + 1, column=1, value="Estimated End Date").font = subheader_font
ws_time.cell(row=summary_row + 1, column=4).value = f"=C{len(time_data)+1}"
ws_time.cell(row=summary_row + 1, column=4).font = Font(name="Calibri", bold=True, color=BLUE, size=12)
ws_time.cell(row=summary_row + 1, column=4).fill = formula_fill
ws_time.cell(row=summary_row + 1, column=4).border = thin_border
ws_time.cell(row=summary_row + 1, column=4).number_format = date_fmt

set_col_widths(ws_time, {"A": 28, "B": 14, "C": 14, "D": 16, "E": 22, "F": 14, "G": 14})
ws_time.freeze_panes = "A2"


# ───────────────────────────────────────────────────────────
# TAB 4: Cost Calculator
# ───────────────────────────────────────────────────────────
ws_cost = wb.create_sheet("Cost Calculator")
ws_cost.sheet_properties.tabColor = ACCENT_YELLOW

# ── INPUT SECTION ──
style_section_header(ws_cost, 1, 1, 4, "PRICING INPUTS", NAVY)

input_labels = [
    ("Hourly Rate ($)", 95, currency_fmt),
    ("Total Hours (from Features)", "='Feature Breakdown'!C15", "0.0"),
    ("Tax Rate (%)", 0.10, pct_fmt),
    ("Platform Fee (%)", 0.05, pct_fmt),
    ("Desired Profit Margin (%)", 0.20, pct_fmt),
]

for i, (label, default, fmt) in enumerate(input_labels):
    r = 3 + i
    ws_cost.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws_cost.cell(row=r, column=1).alignment = Alignment(indent=1)
    cell = ws_cost.cell(row=r, column=3, value=default)
    cell.number_format = fmt
    cell.font = Font(name="Calibri", bold=True, size=12, color=BLUE)
    if isinstance(default, str) and default.startswith("="):
        cell.fill = formula_fill
    else:
        cell.fill = input_fill
        cell.border = input_border

# ── PRICING MODEL SECTION ──
style_section_header(ws_cost, 10, 1, 4, "PRICING MODELS", ACCENT_GREEN)

pricing_rows = [
    ("Hourly Model", "=C3*C4", "Hours x Rate"),
    ("Fixed Price Model", "=C3*C4*(1+C7)", "Hourly + Profit Margin"),
    ("Value-Based Model", "=C3*C4*(1+C7)*1.5", "Fixed + 50% value premium"),
]

ws_cost.cell(row=11, column=1, value="Model").font = subheader_font
ws_cost.cell(row=11, column=2, value="Total Price").font = subheader_font
ws_cost.cell(row=11, column=3, value="Description").font = subheader_font
for c in range(1, 4):
    ws_cost.cell(row=11, column=c).fill = subheader_fill
    ws_cost.cell(row=11, column=c).border = thin_border

for i, (model, formula, desc) in enumerate(pricing_rows):
    r = 12 + i
    ws_cost.cell(row=r, column=1, value=model).font = body_font
    ws_cost.cell(row=r, column=2, value=formula).number_format = currency_fmt
    ws_cost.cell(row=r, column=2).fill = formula_fill
    ws_cost.cell(row=r, column=2).font = Font(name="Calibri", bold=True, size=12, color=ACCENT_GREEN)
    ws_cost.cell(row=r, column=3, value=desc).font = Font(name="Calibri", color=MED_GRAY, size=10)
    style_data_row(ws_cost, r, 3, alt=(i % 2 == 0))

# ── COST BREAKDOWN ──
style_section_header(ws_cost, 17, 1, 4, "COST BREAKDOWN (FIXED PRICE)", BLUE)

breakdown = [
    ("Subtotal (Hours x Rate)", "=C3*C4"),
    ("Profit Margin", "=C18*C7"),
    ("Subtotal + Margin", "=C18+C19"),
    ("Taxes", "=C20*C5"),
    ("Platform Fees", "=C20*C6"),
    ("TOTAL PROJECT COST", "=C20+C21+C22"),
]

for i, (label, formula) in enumerate(breakdown):
    r = 18 + i
    ws_cost.cell(row=r, column=1, value=label).font = Font(
        name="Calibri", bold=(i == len(breakdown) - 1), color=NAVY, size=11
    )
    ws_cost.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_cost.cell(row=r, column=3, value=formula).number_format = currency_fmt
    ws_cost.cell(row=r, column=3).fill = formula_fill
    ws_cost.cell(row=r, column=3).border = thin_border
    if i == len(breakdown) - 1:
        ws_cost.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=14, color=ACCENT_GREEN)
        ws_cost.cell(row=r, column=1).font = Font(name="Calibri", bold=True, color=NAVY, size=12)

# ── WHAT-IF SCENARIOS ──
style_section_header(ws_cost, 26, 1, 4, "WHAT-IF SCENARIOS", PURPLE)

scenario_headers = ["Scenario", "Multiplier", "Hours", "Total Cost"]
for col, h in enumerate(scenario_headers, 1):
    ws_cost.cell(row=27, column=col, value=h)
style_header_row(ws_cost, 27, len(scenario_headers))

scenarios = [
    ("Best Case (optimistic)", 0.8),
    ("Expected Case", 1.0),
    ("Worst Case (pessimistic)", 1.5),
]

for i, (name, mult) in enumerate(scenarios):
    r = 28 + i
    ws_cost.cell(row=r, column=1, value=name).font = body_font
    ws_cost.cell(row=r, column=2, value=mult).number_format = "0.0x"
    ws_cost.cell(row=r, column=3).value = f"=C4*B{r}"
    ws_cost.cell(row=r, column=3).number_format = "0.0"
    ws_cost.cell(row=r, column=3).fill = formula_fill
    ws_cost.cell(row=r, column=4).value = f"=C3*C{r}*(1+C7)"
    ws_cost.cell(row=r, column=4).number_format = currency_fmt
    ws_cost.cell(row=r, column=4).fill = formula_fill
    ws_cost.cell(row=r, column=4).font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    style_data_row(ws_cost, r, 4, alt=(i % 2 == 0))
    # Re-apply formula fills after style_data_row
    ws_cost.cell(row=r, column=3).fill = formula_fill
    ws_cost.cell(row=r, column=4).fill = formula_fill

# Highlight expected case
ws_cost.conditional_formatting.add(
    "A29:D29",
    FormulaRule(formula=["TRUE"],
                fill=PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid"))
)

set_col_widths(ws_cost, {"A": 32, "B": 14, "C": 18, "D": 18})


# ───────────────────────────────────────────────────────────
# TAB 5: Risk Register
# ───────────────────────────────────────────────────────────
ws_risk = wb.create_sheet("Risk Register")
ws_risk.sheet_properties.tabColor = ACCENT_RED

risk_headers = [
    "Risk Description", "Probability", "Impact",
    "Risk Score", "Mitigation Strategy", "Owner", "Status"
]
for col, h in enumerate(risk_headers, 1):
    ws_risk.cell(row=1, column=col, value=h)
style_header_row(ws_risk, 1, len(risk_headers))

# Data validation: Probability and Impact
dv_prob = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
dv_prob.prompt = "Select probability level"
ws_risk.add_data_validation(dv_prob)
dv_prob.add("B2:B50")

dv_impact = DataValidation(type="list", formula1='"Low,Medium,High"', allow_blank=True)
dv_impact.prompt = "Select impact level"
ws_risk.add_data_validation(dv_impact)
dv_impact.add("C2:C50")

dv_risk_status = DataValidation(type="list", formula1='"Open,Mitigated,Closed,Accepted"', allow_blank=True)
ws_risk.add_data_validation(dv_risk_status)
dv_risk_status.add("G2:G50")

# Pre-populated risks
risk_data = [
    ["Scope creep - client adds features mid-project", "High", "High", None,
     "Define clear scope in contract; use change request process", "PM", "Open"],
    ["Key team member unavailable", "Medium", "High", None,
     "Cross-train team; document all decisions", "PM", "Open"],
    ["Third-party API changes or goes down", "Medium", "Medium", None,
     "Use abstraction layer; have fallback plan", "Lead Dev", "Open"],
    ["Client delays feedback/approvals", "High", "Medium", None,
     "Set SLA for feedback in contract; schedule regular reviews", "PM", "Open"],
    ["Underestimated task complexity", "High", "High", None,
     "Add 20% buffer; break tasks into smaller pieces", "Lead Dev", "Open"],
    ["Security vulnerability discovered", "Low", "High", None,
     "Follow OWASP guidelines; schedule security review", "Security", "Open"],
    ["Performance issues at scale", "Medium", "Medium", None,
     "Load test early; design for scalability from start", "Backend Dev", "Open"],
    ["Technology/framework learning curve", "Medium", "Low", None,
     "Allocate training time; use proven tech stack", "Lead Dev", "Open"],
    ["Deployment environment issues", "Low", "Medium", None,
     "Use staging environment; automate deployments", "DevOps", "Open"],
    ["Budget overrun", "Medium", "High", None,
     "Track hours weekly; flag at 75% budget consumed", "PM", "Open"],
]

for i, row_data in enumerate(risk_data, 2):
    for j, val in enumerate(row_data, 1):
        if j != 4:  # Skip risk score (it's a formula)
            ws_risk.cell(row=i, column=j, value=val)

    # Risk Score formula: probability x impact matrix (Low=1, Med=2, High=3)
    ws_risk.cell(row=i, column=4).value = (
        f'=IF(OR(B{i}="",C{i}=""),"",'
        f'(IF(B{i}="Low",1,IF(B{i}="Medium",2,3)))'
        f'*(IF(C{i}="Low",1,IF(C{i}="Medium",2,3))))'
    )
    ws_risk.cell(row=i, column=4).fill = formula_fill
    ws_risk.cell(row=i, column=4).font = Font(name="Calibri", bold=True, size=12)
    ws_risk.cell(row=i, column=4).alignment = Alignment(horizontal="center")

    style_data_row(ws_risk, i, len(risk_headers), alt=(i % 2 == 0))
    # Re-apply formula fill
    ws_risk.cell(row=i, column=4).fill = formula_fill

# Risk score conditional formatting (1-2 green, 3-4 yellow, 6-9 red)
ws_risk.conditional_formatting.add(
    "D2:D50",
    CellIsRule(operator="lessThanOrEqual", formula=["2"],
              fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
              font=Font(color=ACCENT_GREEN, bold=True))
)
ws_risk.conditional_formatting.add(
    "D2:D50",
    CellIsRule(operator="between", formula=["3", "4"],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
              font=Font(color=ACCENT_YELLOW, bold=True))
)
ws_risk.conditional_formatting.add(
    "D2:D50",
    CellIsRule(operator="greaterThanOrEqual", formula=["6"],
              fill=PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid"),
              font=Font(color=ACCENT_RED, bold=True))
)

# Status formatting
ws_risk.conditional_formatting.add(
    "G2:G50",
    CellIsRule(operator="equal", formula=['"Closed"'],
              fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"))
)
ws_risk.conditional_formatting.add(
    "G2:G50",
    CellIsRule(operator="equal", formula=['"Mitigated"'],
              fill=PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid"))
)
ws_risk.conditional_formatting.add(
    "G2:G50",
    CellIsRule(operator="equal", formula=['"Accepted"'],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"))
)

# Risk matrix legend
legend_row = len(risk_data) + 3
ws_risk.cell(row=legend_row, column=1, value="RISK SCORE MATRIX").font = Font(
    name="Calibri", bold=True, color=NAVY, size=12
)
# Matrix headers
for j, lbl in enumerate(["", "Low Impact (1)", "Med Impact (2)", "High Impact (3)"], 1):
    ws_risk.cell(row=legend_row + 1, column=j, value=lbl)
    ws_risk.cell(row=legend_row + 1, column=j).font = header_font
    ws_risk.cell(row=legend_row + 1, column=j).fill = header_fill
    ws_risk.cell(row=legend_row + 1, column=j).alignment = Alignment(horizontal="center")
    ws_risk.cell(row=legend_row + 1, column=j).border = thin_border

for ri, (prob_label, scores) in enumerate([
    ("High Prob (3)", [3, 6, 9]),
    ("Med Prob (2)", [2, 4, 6]),
    ("Low Prob (1)", [1, 2, 3]),
]):
    r = legend_row + 2 + ri
    ws_risk.cell(row=r, column=1, value=prob_label).font = Font(name="Calibri", bold=True, color=NAVY)
    ws_risk.cell(row=r, column=1).fill = subheader_fill
    ws_risk.cell(row=r, column=1).border = thin_border
    for ci, score in enumerate(scores, 2):
        cell = ws_risk.cell(row=r, column=ci, value=score)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Calibri", bold=True, size=14)
        cell.border = thin_border
        if score <= 2:
            cell.fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
        elif score <= 4:
            cell.fill = PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")

set_col_widths(ws_risk, {"A": 40, "B": 14, "C": 14, "D": 12, "E": 45, "F": 14, "G": 12})
ws_risk.freeze_panes = "A2"


# ───────────────────────────────────────────────────────────
# TAB 6: Client Proposal Template
# ───────────────────────────────────────────────────────────
ws_prop = wb.create_sheet("Client Proposal")
ws_prop.sheet_properties.tabColor = PURPLE

# Set print-ready page layout
ws_prop.page_setup.paperSize = ws_prop.PAPERSIZE_LETTER
ws_prop.page_setup.orientation = "portrait"
ws_prop.page_margins.left = 0.75
ws_prop.page_margins.right = 0.75
ws_prop.page_margins.top = 0.75
ws_prop.page_margins.bottom = 0.75

prop_width = 5  # Columns A-E
set_col_widths(ws_prop, {"A": 20, "B": 20, "C": 20, "D": 20, "E": 20})

# ── HEADER ──
ws_prop.merge_cells("A1:E1")
title_cell = ws_prop.cell(row=1, column=1, value="PROJECT PROPOSAL")
title_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=22)
title_cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_prop.row_dimensions[1].height = 50

# Project details from Dashboard
prop_fields = [
    ("Project:", "='Dashboard'!C4", 3),
    ("Client:", "='Dashboard'!C5", 3),
    ("Date:", "='Dashboard'!C6", 3),
    ("Prepared by:", "[Your Name / Company]", 3),
]

for i, (label, value, val_col) in enumerate(prop_fields):
    r = 3 + i
    ws_prop.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws_prop.merge_cells(start_row=r, start_column=val_col, end_row=r, end_column=5)
    cell = ws_prop.cell(row=r, column=val_col, value=value)
    cell.font = Font(name="Calibri", color=DARK_TEXT, size=11)
    if isinstance(value, str) and value.startswith("="):
        cell.fill = formula_fill

# ── SECTION: Project Overview ──
r = 8
style_section_header(ws_prop, r, 1, prop_width, "1. PROJECT OVERVIEW", BLUE)
ws_prop.merge_cells(f"A{r+1}:E{r+3}")
ws_prop.cell(row=r+1, column=1).value = (
    "[Describe the project goals, background, and what the client wants to achieve. "
    "Keep it concise -- 2-3 paragraphs max. Focus on business outcomes, not technical details.]"
)
ws_prop.cell(row=r+1, column=1).font = Font(name="Calibri", color=MED_GRAY, size=10, italic=True)
ws_prop.cell(row=r+1, column=1).alignment = Alignment(wrap_text=True, vertical="top")

# ── SECTION: Scope of Work ──
r = 13
style_section_header(ws_prop, r, 1, prop_width, "2. SCOPE OF WORK", BLUE)

# Pull features from Feature Breakdown
scope_headers = ["Feature", "Priority", "Estimated Hours"]
for col, h in enumerate(scope_headers, 1):
    ws_prop.cell(row=r+1, column=col, value=h)
    ws_prop.cell(row=r+1, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    ws_prop.cell(row=r+1, column=col).fill = blue_accent_fill
    ws_prop.cell(row=r+1, column=col).border = thin_border

# Link to features (first 10 rows)
for fi in range(10):
    fr = r + 2 + fi
    ws_prop.cell(row=fr, column=1).value = f"='Feature Breakdown'!A{fi+2}"
    ws_prop.cell(row=fr, column=1).fill = formula_fill
    ws_prop.cell(row=fr, column=2).value = f"='Feature Breakdown'!D{fi+2}"
    ws_prop.cell(row=fr, column=2).fill = formula_fill
    ws_prop.cell(row=fr, column=3).value = f"='Feature Breakdown'!C{fi+2}"
    ws_prop.cell(row=fr, column=3).fill = formula_fill
    ws_prop.cell(row=fr, column=3).number_format = "0.0"
    for c in range(1, 4):
        ws_prop.cell(row=fr, column=c).border = thin_border
        ws_prop.cell(row=fr, column=c).font = body_font

# ── SECTION: Timeline ──
r = 26
style_section_header(ws_prop, r, 1, prop_width, "3. TIMELINE", BLUE)

tl_headers = ["Phase", "Duration (days)", "Target Dates"]
for col, h in enumerate(tl_headers, 1):
    ws_prop.cell(row=r+1, column=col, value=h)
    ws_prop.cell(row=r+1, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    ws_prop.cell(row=r+1, column=col).fill = blue_accent_fill
    ws_prop.cell(row=r+1, column=col).border = thin_border

for ti in range(7):
    tr = r + 2 + ti
    ws_prop.cell(row=tr, column=1).value = f"='Timeline Builder'!A{ti+2}"
    ws_prop.cell(row=tr, column=1).fill = formula_fill
    ws_prop.cell(row=tr, column=2).value = f"='Timeline Builder'!D{ti+2}"
    ws_prop.cell(row=tr, column=2).fill = formula_fill
    ws_prop.cell(row=tr, column=2).number_format = "0"
    ws_prop.cell(row=tr, column=3).value = f"='Timeline Builder'!B{ti+2}&\" - \"&TEXT('Timeline Builder'!C{ti+2},\"MM/DD/YYYY\")"
    ws_prop.cell(row=tr, column=3).fill = formula_fill
    for c in range(1, 4):
        ws_prop.cell(row=tr, column=c).border = thin_border
        ws_prop.cell(row=tr, column=c).font = body_font

# ── SECTION: Deliverables ──
r = 37
style_section_header(ws_prop, r, 1, prop_width, "4. DELIVERABLES", BLUE)
deliverables = [
    "Complete source code (Git repository access)",
    "Deployment to production environment",
    "Technical documentation and API reference",
    "User guide / admin manual",
    "30 days post-launch bug fix support",
    "[Add/modify deliverables as needed]",
]
for di, item in enumerate(deliverables):
    dr = r + 1 + di
    ws_prop.cell(row=dr, column=1, value=f"  {di+1}. {item}").font = body_font
    ws_prop.merge_cells(f"A{dr}:E{dr}")

# ── SECTION: Pricing ──
r = 45
style_section_header(ws_prop, r, 1, prop_width, "5. PRICING", BLUE)

price_items = [
    ("Total Estimated Hours", "='Cost Calculator'!C4", "0.0"),
    ("Hourly Rate", "='Cost Calculator'!C3", currency_fmt),
    ("Subtotal", "='Cost Calculator'!C18", currency_fmt),
    ("Profit Margin", "='Cost Calculator'!C19", currency_fmt),
    ("Taxes", "='Cost Calculator'!C21", currency_fmt),
    ("TOTAL PROJECT COST", "='Cost Calculator'!C23", currency_fmt),
]

for pi, (label, formula, fmt) in enumerate(price_items):
    pr = r + 1 + pi
    ws_prop.cell(row=pr, column=1, value=label).font = Font(
        name="Calibri", bold=(pi == len(price_items) - 1), color=NAVY, size=11
    )
    ws_prop.cell(row=pr, column=3, value=formula).number_format = fmt
    ws_prop.cell(row=pr, column=3).fill = formula_fill
    ws_prop.cell(row=pr, column=3).border = thin_border
    if pi == len(price_items) - 1:
        ws_prop.cell(row=pr, column=3).font = Font(name="Calibri", bold=True, size=14, color=ACCENT_GREEN)

# ── SECTION: Terms ──
r = 53
style_section_header(ws_prop, r, 1, prop_width, "6. TERMS & CONDITIONS", BLUE)
terms = [
    "Payment: 50% deposit required before work begins, 50% upon completion.",
    "Changes: Scope changes require written change request and may affect timeline/cost.",
    "Timeline: Dates are estimates based on timely client feedback (48hr turnaround).",
    "IP Rights: All intellectual property transfers to client upon final payment.",
    "Cancellation: Either party may cancel with 14 days written notice.",
    "Confidentiality: Both parties agree to keep project details confidential.",
    "[Modify these terms to match your standard contract language.]",
]
for ti, term in enumerate(terms):
    tr = r + 1 + ti
    ws_prop.cell(row=tr, column=1, value=f"  {ti+1}. {term}").font = Font(
        name="Calibri", color=DARK_TEXT if ti < len(terms) - 1 else MED_GRAY,
        size=10, italic=(ti == len(terms) - 1)
    )
    ws_prop.merge_cells(f"A{tr}:E{tr}")


# ───────────────────────────────────────────────────────────
# TAB 1: Dashboard (created last, references all other tabs)
# ───────────────────────────────────────────────────────────
ws_dash = wb.create_sheet("Dashboard", 0)
ws_dash.sheet_properties.tabColor = NAVY

# Title
ws_dash.merge_cells("A1:H1")
title = ws_dash.cell(row=1, column=1, value="PROJECT ESTIMATION DASHBOARD")
title.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
title.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 45

# Instructions
ws_dash.merge_cells("A2:H2")
ws_dash.cell(row=2, column=1, value="Fill in the yellow fields below. All other values update automatically.").font = Font(
    name="Calibri", color=MED_GRAY, size=10, italic=True
)
ws_dash.cell(row=2, column=1).alignment = Alignment(horizontal="center")

# ── PROJECT INFO (input fields) ──
style_section_header(ws_dash, 3, 1, 4, "PROJECT INFORMATION", BLUE)

info_fields = [
    ("Project Name", "E-Commerce Web Application"),
    ("Client Name", "Acme Corp"),
    ("Start Date", "2026-04-14"),
    ("Status", "Not Started"),
]

dv_proj_status = DataValidation(
    type="list", formula1='"Not Started,In Progress,Complete"', allow_blank=True
)
ws_dash.add_data_validation(dv_proj_status)
dv_proj_status.add("C7")

for i, (label, default) in enumerate(info_fields):
    r = 4 + i
    ws_dash.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, color=NAVY, size=11)
    ws_dash.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_dash.merge_cells(f"C{r}:D{r}")
    cell = ws_dash.cell(row=r, column=3, value=default)
    cell.font = Font(name="Calibri", bold=True, size=12, color=BLUE)
    cell.fill = input_fill
    cell.border = input_border
    if label == "Start Date":
        cell.number_format = date_fmt

# Status conditional formatting
ws_dash.conditional_formatting.add(
    "C7:D7",
    CellIsRule(operator="equal", formula=['"Not Started"'],
              fill=PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid"))
)
ws_dash.conditional_formatting.add(
    "C7:D7",
    CellIsRule(operator="equal", formula=['"In Progress"'],
              fill=PatternFill(start_color="FEFCBF", end_color="FEFCBF", fill_type="solid"),
              font=Font(color=ACCENT_YELLOW, bold=True))
)
ws_dash.conditional_formatting.add(
    "C7:D7",
    CellIsRule(operator="equal", formula=['"Complete"'],
              fill=PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
              font=Font(color=ACCENT_GREEN, bold=True))
)

# ── KEY METRICS ──
style_section_header(ws_dash, 9, 1, 8, "KEY METRICS", ACCENT_GREEN)

# Row of metric boxes
metrics = [
    ("Total Features", "=COUNTA('Feature Breakdown'!A2:A100)", None),
    ("Total Hours", "='Feature Breakdown'!C15", None),
    ("Hourly Rate", "='Cost Calculator'!C3", currency_fmt),
    ("Project Cost", "='Cost Calculator'!C23", currency_fmt),
    ("Timeline (days)", "='Timeline Builder'!D10", None),
    ("Est. End Date", "='Timeline Builder'!D11", date_fmt),
    ("High Risks", "=COUNTIF('Risk Register'!D2:D50,\">=\"&6)", None),
    ("Completion %", "='Feature Breakdown'!C17", pct_fmt),
]

for mi, (label, formula, fmt) in enumerate(metrics):
    col = mi + 1
    style_metric_box(ws_dash, 10, col, label, formula, fmt)

# ── RISK-ADJUSTED TIMELINE ──
style_section_header(ws_dash, 13, 1, 8, "RISK-ADJUSTED ESTIMATES", ACCENT_ORANGE)

adj_labels = [
    ("Base Timeline (days)", "='Timeline Builder'!D10", "0"),
    ("Risk Buffer (+20%)", "='Timeline Builder'!D10*0.2", "0.0"),
    ("Risk-Adjusted Timeline", "='Timeline Builder'!D10*1.2", "0"),
    ("Risk-Adjusted End Date", "=C6+'Timeline Builder'!D10*1.2", date_fmt),
]

for i, (label, formula, fmt) in enumerate(adj_labels):
    r = 14 + i
    ws_dash.cell(row=r, column=1, value=label).font = Font(
        name="Calibri", bold=(i >= 2), color=NAVY, size=11
    )
    ws_dash.cell(row=r, column=1).alignment = Alignment(indent=1)
    ws_dash.cell(row=r, column=3, value=formula).number_format = fmt
    ws_dash.cell(row=r, column=3).fill = formula_fill
    ws_dash.cell(row=r, column=3).border = thin_border
    if i >= 2:
        ws_dash.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=12, color=ACCENT_ORANGE)

# ── WHAT-IF COST SUMMARY ──
style_section_header(ws_dash, 19, 1, 8, "COST SCENARIOS", PURPLE)

scenario_labels = [
    ("Best Case (0.8x)", "='Cost Calculator'!D28", currency_fmt),
    ("Expected (1.0x)", "='Cost Calculator'!D29", currency_fmt),
    ("Worst Case (1.5x)", "='Cost Calculator'!D30", currency_fmt),
]

for i, (label, formula, fmt) in enumerate(scenario_labels):
    col = 1 + i * 3
    ws_dash.merge_cells(start_row=20, start_column=col, end_row=20, end_column=col + 1)
    ws_dash.merge_cells(start_row=21, start_column=col, end_row=21, end_column=col + 1)
    style_metric_box(ws_dash, 20, col, label, formula, fmt)

# ── PRIORITY BREAKDOWN ──
style_section_header(ws_dash, 23, 1, 8, "FEATURE PRIORITY BREAKDOWN", BLUE)

prio_headers = ["Priority", "Count", "Hours"]
for col, h in enumerate(prio_headers, 1):
    ws_dash.cell(row=24, column=col, value=h)
style_header_row(ws_dash, 24, 3)

priorities = ["Must Have", "Should Have", "Could Have", "Won't Have"]
for pi, prio in enumerate(priorities):
    pr = 25 + pi
    ws_dash.cell(row=pr, column=1, value=prio).font = body_font
    ws_dash.cell(row=pr, column=2).value = f'=COUNTIF(\'Feature Breakdown\'!D2:D100,"{prio}")'
    ws_dash.cell(row=pr, column=2).fill = formula_fill
    ws_dash.cell(row=pr, column=2).alignment = Alignment(horizontal="center")
    ws_dash.cell(row=pr, column=3).value = f'=SUMIF(\'Feature Breakdown\'!D2:D100,"{prio}",\'Feature Breakdown\'!C2:C100)'
    ws_dash.cell(row=pr, column=3).fill = formula_fill
    ws_dash.cell(row=pr, column=3).number_format = "0.0"
    style_data_row(ws_dash, pr, 3, alt=(pi % 2 == 0))
    ws_dash.cell(row=pr, column=2).fill = formula_fill
    ws_dash.cell(row=pr, column=3).fill = formula_fill

# Priority bar chart
chart = BarChart()
chart.type = "col"
chart.title = "Hours by Priority"
chart.y_axis.title = "Hours"
chart.x_axis.title = "Priority Level"
chart.style = 10

cats = Reference(ws_dash, min_col=1, min_row=25, max_row=28)
vals = Reference(ws_dash, min_col=3, min_row=24, max_row=28)
chart.add_data(vals, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 22
chart.height = 12

ws_dash.add_chart(chart, "E23")

set_col_widths(ws_dash, {
    "A": 22, "B": 14, "C": 18, "D": 14,
    "E": 14, "F": 14, "G": 14, "H": 14
})


# ═══════════════════════════════════════════════════════════
# SAVE WORKBOOK
# ═══════════════════════════════════════════════════════════
output_dir = os.path.dirname(os.path.abspath(__file__))
filename = "Developer-Project-Estimator.xlsx"
filepath = os.path.join(output_dir, filename)
wb.save(filepath)
print(f"Workbook saved: {filepath}")
print(f"Tabs: {[ws.title for ws in wb.worksheets]}")
